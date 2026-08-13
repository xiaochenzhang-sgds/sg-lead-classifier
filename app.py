"""
Sales Ops · Data Quality Suite — Singapore
Delivery Hero / Pandora · Digital Sales APAC

TAB 1  Lead Classification
       Postal + Name dedup (unit used as filter)
       Labels: P1 / P2 / P3 / P4 / Business Closed / Wrong Target Group

TAB 2  Generate Apify URLs

TAB 3  SF Account Audit
       Find suspected duplicate pairs within Salesforce itself

Label reference:
  P1  — New               No CRM match, Apify-confirmed restaurant
  P2  — Please Check      No Apify result or no category found
  P3  — Potential Match   Name 0.50–0.74 at same postal code
  P4  — Duplicate         Name ≥ 0.75 at same postal code
  Business Closed         Apify: Google confirms permanently/temporarily closed
  Wrong Target Group      Apify: category not food-delivery eligible
"""

import streamlit as st
import pandas as pd
import re
import io
import time
import json
from urllib.parse import unquote, quote, urlencode
from urllib.request import urlopen
from difflib import SequenceMatcher          # used only for street_similarity
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from shapely.wkt import loads as wkt_loads
from shapely.geometry import Point
from rapidfuzz import fuzz

try:
    from pypinyin import lazy_pinyin, Style as PinyinStyle
    _PYPINYIN_AVAILABLE = True
except ImportError:
    _PYPINYIN_AVAILABLE = False


# ── Secrets & caches ─────────────────────────────────────────────
APP_PASSWORD        = st.secrets["APP_PASSWORD"]
ONEMAP_CREDENTIALS  = {"email": st.secrets["ONEMAP_EMAIL"],
                        "password": st.secrets["ONEMAP_PASSWORD"]}
_TOKEN_CACHE        = {"token": None, "expiry": 0}


def _load_logo():
    import os, base64
    p = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dh_logo.png")
    return base64.b64encode(open(p, "rb").read()).decode() if os.path.exists(p) else ""

DH_LOGO_B64 = _load_logo()


def check_password():
    if st.session_state.get("authenticated"):
        return True
    _, col, _ = st.columns([1, 1.6, 1])
    with col:
        st.markdown(f'''
        <div style="text-align:center;padding:2rem 0 1rem 0;">
            <img src="data:image/png;base64,{DH_LOGO_B64}"
                 style="width:140px;margin-bottom:1.2rem;" />
            <h2 style="color:#1A1A1A;font-size:1.4rem;font-weight:700;margin-bottom:0.2rem;">
                Sales Ops · Data Quality Suite</h2>
            <p style="color:#888;font-size:0.88rem;margin-bottom:1.4rem;">
                Digital Sales APAC · Pandora / Delivery Hero</p>
        </div>''', unsafe_allow_html=True)
        pwd = st.text_input("Password", type="password",
                            placeholder="Enter password to continue")
        if st.button("Sign in", type="primary", use_container_width=True):
            if pwd == APP_PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    return False


st.set_page_config(page_title="Sales Ops Suite",
                   page_icon="🎯", layout="wide",
                   initial_sidebar_state="expanded")


# ═════════════════════════════════════════════════════════════════
# STATUS GROUPS  (shared by both tabs)
# ═════════════════════════════════════════════════════════════════

ACTIVE_PIPELINE = [
    "active", "new", "collecting documents", "negotiation",
    "menu processing", "onboarding", "quality check",
]
WIN_BACK        = ["lost", "terminated"]
WIN_BACK_FAILED = ["win back failed"]


def has_closed_marker(name) -> bool:
    """True if the account name ends with '- Closed' (common SF convention)."""
    if not name or pd.isna(name):
        return False
    return bool(re.search(r'-\s*closed', str(name).strip().lower()))


def get_status_group(status, is_closed_name: bool = False) -> str:
    if is_closed_name:
        return "win_back"
    s = str(status).strip().lower()
    if s in WIN_BACK_FAILED: return "win_back_failed"
    if s in WIN_BACK:        return "win_back"
    return "active_pipeline"


def get_crm_label(match_method: str, match_score: float,
                  status_group: str, p2: float, p3: float) -> str:
    """Map (match quality, CRM status) → lead priority label."""
    if status_group == "win_back_failed":
        return "WBF — Win-back Failed"
    if status_group == "win_back":
        return "WB — Win-back"
    # active_pipeline
    if match_method == "GRID Exact" or match_score >= 1.0:
        return "P4 — Duplicate (Exact)"
    if match_score >= p3:
        return "P3 — Duplicate"
    if match_score >= p2:
        return "P2 — Potential"
    return "Needs Review"


def get_risk_and_action(status_a, status_b):
    """Risk level + recommended action for SF Account Audit pairs."""
    def grp(s):
        sl = str(s).strip().lower()
        if sl in WIN_BACK_FAILED: return "wbf"
        if sl in WIN_BACK:        return "winback"
        return "pipeline"
    ga, gb = grp(status_a), grp(status_b)
    pair   = frozenset([ga, gb])
    if pair == frozenset(["pipeline"]):
        if (str(status_a).strip().lower() == "active"
                and str(status_b).strip().lower() == "active"):
            return "🔴 High",   "Merge required — two live active accounts"
        return     "🔴 High",   "Being worked twice — check with rep"
    if pair == frozenset(["pipeline", "winback"]):
        return     "🟡 Medium", "Review — possibly stale record from prior attempt"
    if pair == frozenset(["winback"]):
        return     "🟢 Low",    "Low priority — both inactive"
    if "wbf" in pair:
        return     "🟢 Low",    "Informational — win-back already attempted"
    return         "🟡 Medium", "Review recommended"


# ═════════════════════════════════════════════════════════════════
# MARKET CONFIG
# ═════════════════════════════════════════════════════════════════

MARKETS = {
    "SG": {
        "code": "SG", "name": "Singapore", "flag": "🇸🇬",
        "char_map": {}, "country_suffix": "Singapore", "phone_prefix": "65",
    }
}


# ═════════════════════════════════════════════════════════════════
# FOOD DELIVERY ELIGIBILITY
# ═════════════════════════════════════════════════════════════════

FOOD_DELIVERY_ALLOWED = {
    "Restaurant","Fine dining restaurant","Family restaurant","Casual dining restaurant",
    "Buffet restaurant","Brasserie","Bistro","Diner","Eatery","Pizza restaurant",
    "Pizza delivery","Pizza takeaway","Kebab shop","Kebab restaurant",
    "Doner kebab restaurant","Shawarma restaurant","Falafel restaurant",
    "Pita restaurant","Sushi restaurant","Ramen restaurant","Noodle restaurant",
    "Dumpling restaurant","Dim sum restaurant","Wonton restaurant","Steak house",
    "Steakhouse","Grill restaurant","Barbecue restaurant","BBQ restaurant",
    "Smokehouse","Rotisserie chicken restaurant","Sandwich shop",
    "Submarine sandwich shop","Wrap restaurant","Salad shop","Bowl restaurant",
    "Poke bar","Soup restaurant","Soup kitchen","Breakfast restaurant",
    "Brunch restaurant","Pancake restaurant","Waffle restaurant",
    "Dessert restaurant","Dessert shop","Ice cream shop","Ice cream parlor",
    "Frozen yogurt shop","Donut shop","Doughnut shop","Crepe restaurant",
    "Waffle house","Vegetarian restaurant","Vegan restaurant",
    "Plant-based restaurant","Organic restaurant","Health food restaurant",
    "Gluten-free restaurant","Halal restaurant","Kosher restaurant","Food hall",
    "Food truck","Street food restaurant","Market restaurant",
    "Home cooking restaurant","Traditional restaurant","Local restaurant",
    "Neighborhood restaurant","Deli","Delicatessen","Charcuterie","Lunchroom",
    "Snack bar","Juice bar","Smoothie bar","Açaí shop","Chocolate shop",
    "Sweet shop","Candy store","Noodle shop","Pasta shop","Rice restaurant",
    "Porridge restaurant","Congee restaurant","Hot pot restaurant",
    "Fondue restaurant","Raclette restaurant","Teppanyaki restaurant",
    "Okonomiyaki restaurant","Takoyaki restaurant","Yakitori restaurant",
    "Izakaya","Robatayaki restaurant","Tempura restaurant","Tonkatsu restaurant",
    "Udon restaurant","Soba restaurant","Gyoza restaurant","Pho restaurant",
    "Banh mi restaurant","Spring roll restaurant","Satay restaurant",
    "Rendang restaurant","Curry restaurant","Tandoori restaurant",
    "Biryani restaurant","Dosa restaurant","Idli restaurant","Thali restaurant",
    "Ceviche restaurant","Empanada restaurant","Arepas restaurant",
    "Chimichanga restaurant","Meal delivery","Food delivery","Takeaway",
    "Takeout restaurant","Take-out restaurant","Cloud kitchen","Ghost kitchen",
    "Virtual restaurant",
    # SG-specific
    "Zi char restaurant","Cze char restaurant","Tze char restaurant",
    "Economy rice stall","Nasi lemak restaurant","Chicken rice restaurant",
    "Laksa restaurant","Wonton mee restaurant","Char kway teow restaurant",
    "Bak kut teh restaurant","Murtabak restaurant","Prata restaurant",
    "Mixed rice restaurant","Hawker-style restaurant","Nasi padang restaurant",
    "Hawker stall",
    # Café & bakery
    "Cafe","Coffee shop","Coffee house","Coffeehouse","Espresso bar",
    "Tea house","Bubble tea shop","Boba shop","Bakery","Patisserie",
    "Pastry shop","Cake shop","Bread bakery","Artisan bakery","French bakery",
    "Cookie shop","Cupcake shop","Bagel shop",
    # Fast food / QSR
    "Fast food restaurant","Fast-food restaurant","Quick service restaurant",
    "Hamburger restaurant","Burger restaurant","Hot dog restaurant",
    "Fried chicken restaurant","Chicken restaurant","Chicken wings restaurant",
    "Fish and chips restaurant","Seafood restaurant","Fish restaurant",
    "Taco restaurant","Burrito restaurant",
    # Cuisines
    "Italian restaurant","French restaurant","Chinese restaurant",
    "Japanese restaurant","Thai restaurant","Indian restaurant",
    "Mexican restaurant","Greek restaurant","Lebanese restaurant",
    "Middle Eastern restaurant","Mediterranean restaurant","Asian restaurant",
    "Korean restaurant","Vietnamese restaurant","Spanish restaurant",
    "American restaurant","Ethiopian restaurant","Afghan restaurant",
    "Pakistani restaurant","Nepalese restaurant","Sri Lankan restaurant",
    "Bangladeshi restaurant","Indonesian restaurant","Filipino restaurant",
    "Peruvian restaurant","Brazilian restaurant","Argentinian restaurant",
    "Georgian restaurant","Uzbek restaurant","Syrian restaurant",
    "Moroccan restaurant","Egyptian restaurant","Caribbean restaurant",
    "Jamaican restaurant","Cuban restaurant","Portuguese restaurant",
    "German restaurant","Austrian restaurant","Scandinavian restaurant",
    "Nordic restaurant","Latin American restaurant","Fusion restaurant",
    "International restaurant","European restaurant","Pan-Asian restaurant",
    "Oriental restaurant",
    # SG dialect / local tags
    "Hokkien","Teochew","Cantonese","Dim Sum","Hainanese","Hakka","Shanghai",
    "Sichuan","Hunan","Jiang Su","Putian","Dong Bei",
    "Hong Kong (Cha Chaan Teng)","Seafood BBQ","Western","Dessert","Tang Shui",
    "Cakes","Malay","Indonesian","Indian","Nyonya","Peranakan","Taiwan",
    "Taiwanese","Turkish","Roast Meat","Nasi Padang","Indian Muslim",
    "South Indian","Eurasian","Hotpot","Buffet","Seafood","Fusion","British",
    "Australian","Cajun","Caribbean","Cuban","Greek","Halal","Internation",
    "Mala","Mookata","Muslim","Middle Eastern","Steakhouse","Swedish","Vegan",
    "Izakaya","Local",
}

_DEFAULT_EXCLUSION_KW = [
    "hawker centre","hawker center","kopitiam",
    "food court","food centre","food center","eating house",
    "club","hotel","cantine","canteen","brewery","liquor","wine",
    "酒吧","酒店","美食广场","美食中心","食阁","咖啡店","巴刹",
    "karaoke","卡拉ok","bbq space","event space","salon",
    "coffee beans retailer","shopping mall","catering","roastery",
    "vending machine","cyber cafe","grocery","supermarket",
    "alcohol shop","fruit shop","events","food festival","pop-up",
    "night club","pet bakery","酒馆","沙龙","自动贩卖机","商场","超市",
    "cc restaurant","community club","army","saf","mindef",
    "school canteen","hospital canteen","industrial canteen",
    "mix & match",
]


def is_food_delivery_eligible(category, exclusion_kw: list) -> bool:
    if not category or pd.isna(category):
        return False
    cat = str(category).strip()
    cat_lower = cat.lower()
    if any(kw in cat_lower for kw in exclusion_kw):
        return False
    if cat in FOOD_DELIVERY_ALLOWED:
        return True
    for allowed in FOOD_DELIVERY_ALLOWED:
        if allowed.lower() == cat_lower:
            return True
    food_kw = [
        "restaurant","cafe","café","bakery","kebab","pizza","sushi","burger",
        "grill","bistro","brasserie","diner","eatery","kitchen","takeaway",
        "takeout","delivery","patisserie","pastry","coffee","tea house",
        "noodle","ramen","deli","snack","food truck","steakhouse","seafood",
        "sandwich","dim sum","hotpot","steak house","餐厅","餐馆","小吃","烘焙","面包",
    ]
    return any(kw in cat_lower for kw in food_kw)


# ═════════════════════════════════════════════════════════════════
# SINGAPORE UTILITY FUNCTIONS
# ═════════════════════════════════════════════════════════════════

SG_UNIT_RE   = re.compile(r'#\s*(\d{1,2})\s*[-–]\s*(\d{1,4}[A-Za-z]?)', re.IGNORECASE)
SG_POSTAL_RE = re.compile(r'\b(0[1-9]|[1-7]\d|8[0-2])\d{4}\b')
SG_NAME_NOISE= re.compile(
    r'\b(pte\.?\s*ltd\.?|sdn\.?\s*bhd\.?|llp|incorporated|'
    r'holdings?|group|enterprise[s]?|trading|singapore|sg)\b',
    re.IGNORECASE)

# Common Singapore area names for geographic conflict detection
# in the zero-postal name-only fallback.
SG_AREAS = {
    "tampines","jurong","woodlands","yishun","hougang","sengkang","punggol",
    "bishan","ang mo kio","bedok","pasir ris","changi","geylang","kallang",
    "clementi","choa chu kang","sembawang","queenstown","toa payoh",
    "buona vista","orchard","serangoon","bukit timah","bukit batok",
    "paya lebar","novena","marine parade","marsiling","simei","boon lay",
    "admiralty","bras basah","chinatown","dhoby ghaut","farrer park",
    "harbourfront","hillview","jurong east","jurong west","kovan","lavender",
    "little india","macpherson","newton","potong pasir","redhill","tiong bahru",
    "upper thomson","vivocity","whampoa",
}

# ── Strip venue names and generic words before name comparison ────
# Parenthetical content (e.g. "(Tanglin Mall)", "(White Sands)") and
# generic words ("restaurant/s") inflate fuzzy scores between
# co-located but unrelated businesses. Strip before normalising.
_PAREN_RE   = re.compile(r'\([^)]*\)', re.UNICODE)
_GENERIC_RE = re.compile(r'\b(restaurants?)\b', re.IGNORECASE)

def strip_venue_generic(name: str) -> str:
    """Remove parenthetical venue tags and generic restaurant words.

    Examples:
        'Western Grill (Tanglin Mall)'          → 'Western Grill'
        'KFC (White Sands)'                     → 'KFC'
        'DOMO Restaurant'                        → 'DOMO'
        'Shahi Maharani North Indian Restaurant' → 'Shahi Maharani North Indian'
    """
    s = _PAREN_RE.sub(' ', str(name or ''))
    s = _GENERIC_RE.sub(' ', s)
    return re.sub(r'\s+', ' ', s).strip()
_CHINESE_RE  = re.compile(r'[\u4e00-\u9fff\u3400-\u4dbf\uf900-\ufaff]')
_NA_VALUES   = {"","nan","none","n/a","na","nil","-","–",
                "unknown","no name","na/","n.a.","n.a"}


def extract_sg_unit(text: str) -> str:
    if not text or str(text).strip() in ("","nan"):
        return ""
    m = SG_UNIT_RE.search(str(text))
    if m:
        return f"{m.group(1).zfill(2)}-{m.group(2).lower().zfill(3)}"
    return ""


def extract_sg_postal(text: str) -> str:
    if not text or str(text).strip() in ("","nan"):
        return ""
    m = SG_POSTAL_RE.search(str(text))
    return m.group(0) if m else ""


# Placeholder / obviously fake postal codes that should be treated as blank
_FAKE_POSTALS = frozenset({"000000","000001","111111","123456","999999"})

def _norm_postal_input(raw) -> str:
    """
    Normalise a raw postal column value before extraction:
    - Strips non-digits and trailing .0 (Excel float artefact)
    - Zero-pads 5-digit codes (Excel drops leading zero for 09xxxx postals)
    - Returns '' for known fake / placeholder codes

    Only applies to postal COLUMN values, not to full address strings.
    """
    s = re.sub(r'\D', '', str(raw or "").strip().replace(".0",""))
    if len(s) == 5:
        s = "0" + s
    if s in _FAKE_POSTALS:
        return ""
    return s


def fix_postal_sg(postal) -> str:
    """Normalise to 6-digit string for SF Account Audit preprocessing."""
    s = _norm_postal_input(postal)
    return s if s else str(postal).strip().replace(".0","")  # fallback: return as-is


def is_blank_name(s) -> bool:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return True
    return str(s).strip().lower() in _NA_VALUES


def has_chinese(text: str) -> bool:
    return bool(_CHINESE_RE.search(str(text or "")))


def to_pinyin(text: str) -> str:
    """
    Converts Chinese characters to Hanyu Pinyin (no tones).
    Non-Chinese text is passed through as WHOLE SEGMENTS (not char by char)
    to prevent single-letter tokens causing spurious fuzzy matches.

    Examples:
        '海底捞'               → 'hai di lao'
        'ABC 餐厅'             → 'abc can ting'
        'Master Tang 大堂炒饭'  → 'master tang da tang chao fan'
    """
    if not text or is_blank_name(text):
        return ""
    text = str(text).strip()
    if not _PYPINYIN_AVAILABLE:
        return text.lower()
    result  = []
    segment = []                        # buffer for consecutive non-Chinese chars
    for char in text:
        if _CHINESE_RE.match(char):
            if segment:                 # flush non-Chinese buffer as one word group
                result.append("".join(segment))
                segment = []
            result.extend(lazy_pinyin(char, style=PinyinStyle.NORMAL))
        else:
            segment.append(char)
    if segment:
        result.append("".join(segment)) # flush any trailing non-Chinese text
    return " ".join(result).lower().strip()


def norm_name_sg(s, char_map: dict) -> tuple:
    """Returns (latin_norm, pinyin_norm) — strips SG entity noise."""
    if is_blank_name(s):
        return "", ""
    s = str(s).strip()
    for k, v in char_map.items():
        s = s.replace(k, v)
    s = SG_NAME_NOISE.sub("", s)
    s = SG_UNIT_RE.sub("", s)
    s = re.sub(r'@\w+', "", s)
    s = re.sub(r'\s+', " ", s).strip()
    return s.lower().strip(), (to_pinyin(s) if has_chinese(s) else "")


# ═════════════════════════════════════════════════════════════════
# CORE HELPERS
# ═════════════════════════════════════════════════════════════════

def get_onemap_token():
    from urllib.request import Request
    now = time.time()
    if _TOKEN_CACHE["token"] and _TOKEN_CACHE["expiry"] > (now + 3600):
        return _TOKEN_CACHE["token"]
    try:
        payload = json.dumps(ONEMAP_CREDENTIALS).encode()
        req = Request("https://www.onemap.gov.sg/api/auth/post/getToken",
                      data=payload,
                      headers={"Content-Type":"application/json",
                               "User-Agent":"SalesOpsSuite/2.0"},
                      method="POST")
        with urlopen(req, timeout=5) as resp:
            data = json.loads(resp.read().decode())
            if "access_token" in data:
                _TOKEN_CACHE["token"]  = data["access_token"]
                _TOKEN_CACHE["expiry"] = now + 259200
                return data["access_token"]
    except Exception as e:
        st.error(f"OneMap token error: {e}")
    return None


def norm_name(s, char_map: dict) -> str:
    if pd.isna(s): return ""
    s = str(s).strip()
    for k, v in char_map.items():
        s = s.replace(k, v)
    s = SG_NAME_NOISE.sub("", s)
    return re.sub(r'\s+', ' ', s).strip().lower()


def name_confidence(a, b, char_map: dict) -> float:
    """
    Name similarity using rapidfuzz token_sort_ratio (handles word order).
    Also compares Hanyu Pinyin when either string contains Chinese.
    Returns 0.0–1.0.
    """
    a_n = norm_name(a, char_map)
    b_n = norm_name(b, char_map)
    if not a_n or not b_n:
        return 0.0
    # token_sort_ratio: handles word order (A B C == C B A)
    # token_set_ratio:  handles subset matching — critical for CRM names that
    #                   append location in brackets, e.g. "Yummy Taste" vs
    #                   "Yummy Taste (Rivervale Drive)" → 100% not 55%
    ts_score    = fuzz.token_sort_ratio(a_n, b_n) / 100.0
    tset_score  = fuzz.token_set_ratio(a_n, b_n)  / 100.0
    latin_score = max(ts_score, tset_score)
    if has_chinese(str(a or "")) or has_chinese(str(b or "")):
        pin_a = to_pinyin(str(a or ""))
        pin_b = to_pinyin(str(b or ""))
        if pin_a and pin_b:
            p_ts         = fuzz.token_sort_ratio(pin_a, pin_b) / 100.0
            p_tset       = fuzz.token_set_ratio(pin_a, pin_b)  / 100.0
            pinyin_score = max(p_ts, p_tset)
            return round(max(latin_score, pinyin_score), 3)
    return round(latin_score, 3)


def norm_phone(p, prefix: str) -> str:
    if pd.isna(p): return ""
    s = str(p).replace("+","").replace(" ","").replace("-","").strip()
    if s.endswith(".0"): s = s[:-2]
    s = s.lstrip("0")
    if not s.startswith(prefix): s = prefix + s
    local = re.sub(r"\D","",s)[len(prefix):]
    if len(local) != 8 or local[0] not in ("6","8","9"): return ""
    return s


def to_e164(p, prefix: str) -> str:
    n = norm_phone(p, prefix)
    return "+" + n if n else ""


def norm_url(u) -> str:
    if pd.isna(u): return ""
    u = unquote(str(u).strip())
    u = u.split("?hl=")[0]
    u = u.split("&hl=")[0]              # searchPageUrl appends &hl=en
    u = u.split("&query_place_id=")[0]  # strip Apify-appended place ID
    return u.lower()


def address_match(lead_street, apify_address, char_map: dict) -> bool:
    if pd.isna(lead_street) or pd.isna(apify_address): return False
    ls = norm_name(str(lead_street), char_map)
    aa = norm_name(str(apify_address), char_map)
    pl = set(re.findall(r"\b\d{6}\b", ls))
    pa = set(re.findall(r"\b\d{6}\b", aa))
    if pl and pa and pl & pa:
        ul, ua = extract_sg_unit(ls), extract_sg_unit(aa)
        if ul and ua: return ul == ua
        return True
    ignore = {"","no","sk","singapore","blk","block","新加坡","大厦","路","街"}
    tl = set(re.split(r"\W+",ls)) - ignore
    ta = set(re.split(r"\W+",aa)) - ignore
    return len(tl & ta) >= 2


def street_similarity(street_a, street_b):
    def _n(s):
        if not s or str(s).strip() in ("","nan"): return ""
        s = str(s).upper().strip()
        s = re.sub(r'\b\d+[A-Z]?\b','',s)
        s = re.sub(r'\b(BLK|BLOCK)\b','',s)
        return re.sub(r'\s+',' ',s).strip()
    na, nb = _n(street_a), _n(street_b)
    if not na or not nb: return None
    return SequenceMatcher(None, na, nb).ratio()


def detect_column(df: pd.DataFrame, candidates: list):
    for c in candidates:
        if c in df.columns: return c
    lmap = {col.lower(): col for col in df.columns}
    for c in candidates:
        if c.lower() in lmap: return lmap[c.lower()]
    return None


def find_header_row(file, key_col="GRID") -> int:
    try:
        raw = pd.read_excel(file, header=None, nrows=30)
        for i, row in raw.iterrows():
            if any(str(v).strip() == key_col for v in row.dropna()): return i
    except Exception: pass
    return 0


# ═════════════════════════════════════════════════════════════════
# CACHED FILE LOADERS
# ═════════════════════════════════════════════════════════════════

@st.cache_data(show_spinner=False)
def _cached_read(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Read raw bytes → DataFrame, cached so re-uploads don't re-parse.
    Supports .xlsx, .xls (real binary), .xls (HTML Salesforce export), and .csv."""
    from io import StringIO
    if filename.endswith(".xlsx"):
        return pd.read_excel(io.BytesIO(file_bytes))
    if filename.endswith(".xls"):
        # Salesforce often exports HTML tables with an .xls extension.
        # Detect by checking if the file starts with an HTML tag.
        header = file_bytes[:512].lstrip()
        is_html = (header.startswith(b"<") or
                   b"<html" in header.lower() or
                   b"<head" in header.lower() or
                   b"<table" in header.lower())
        if is_html:
            try:
                tables = pd.read_html(io.BytesIO(file_bytes))
                if tables:
                    return tables[0]
            except Exception:
                pass
            # Last resort: tab-separated text (some Salesforce HTML exports)
            for enc in ("utf-8", "utf-8-sig", "windows-1252", "latin-1"):
                try:
                    raw = file_bytes.decode(enc)
                    sep = "\t" if raw.count("\t") > raw.count(",") else ","
                    return pd.read_csv(StringIO(raw), sep=sep,
                                       on_bad_lines="skip", engine="python")
                except Exception:
                    continue
        else:
            try:
                return pd.read_excel(io.BytesIO(file_bytes), engine="xlrd")
            except Exception:
                pass
    for enc in ("utf-8","utf-8-sig","windows-1252","latin-1"):
        try:
            raw = file_bytes.decode(enc)
            sep = ";" if raw.count(";") > raw.count(",") else ","
            return pd.read_csv(StringIO(raw), sep=sep, quotechar='"',
                               on_bad_lines="skip", engine="python")
        except Exception: continue
    raw = file_bytes.decode("latin-1", errors="replace")
    sep = ";" if raw.count(";") > raw.count(",") else ","
    return pd.read_csv(StringIO(raw), sep=sep, quotechar='"',
                       on_bad_lines="skip", engine="python")


@st.cache_data(show_spinner="Indexing CRM accounts — runs once per session…")
def _cached_preprocess_crm(df_raw: pd.DataFrame,
                             name_col, postal_col, addr_col,
                             status_col, char_map_tuple: tuple) -> pd.DataFrame:
    """
    Heavy CRM preprocessing (normalisation, pinyin, closed-name detection).
    Cached: re-uploading the same file skips all processing.
    """
    char_map = dict(char_map_tuple)
    df = df_raw.copy()
    if name_col:
        df["_is_closed_name"] = df[name_col].apply(has_closed_marker)
        df["_name_latin"]     = df[name_col].apply(
            lambda n: norm_name_sg(strip_venue_generic(n), char_map)[0])
        df["_name_pinyin"]    = df[name_col].apply(
            lambda n: norm_name_sg(strip_venue_generic(n), char_map)[1])
    else:
        df["_is_closed_name"] = False
        df["_name_latin"]     = ""
        df["_name_pinyin"]    = ""
    if postal_col:
        df["_postal_fixed"]   = df[postal_col].apply(
            lambda p: extract_sg_postal(_norm_postal_input(p))
                      or fix_postal_sg(p))
    else:
        df["_postal_fixed"]   = ""
    if addr_col:
        df["_unit_extracted"] = df[addr_col].apply(extract_sg_unit)
    else:
        df["_unit_extracted"] = ""
    if status_col:
        df["_status_group"]   = df.apply(
            lambda r: get_status_group(r[status_col], r["_is_closed_name"]), axis=1)
    else:
        df["_status_group"]   = "active_pipeline"
    return df


def load_leads(file_bytes: bytes, filename: str, market_cfg: dict):
    df = _cached_read(file_bytes, filename)
    if filename.endswith((".xlsx", ".xls")):
        # Try to detect header row for Excel files with GRID column
        from io import BytesIO
        engine = "xlrd" if filename.endswith(".xls") else None
        try:
            raw_xl = pd.read_excel(BytesIO(file_bytes), header=None, nrows=30,
                                   **({"engine": engine} if engine else {}))
            for i, row in raw_xl.iterrows():
                if any(str(v).strip() == "GRID" for v in row.dropna()):
                    df = pd.read_excel(BytesIO(file_bytes), header=i,
                                       **({"engine": engine} if engine else {}))
                    break
        except Exception: pass
        grid_col = next((c for c in df.columns if str(c).strip() == "GRID"), None)
        if grid_col:
            df = df[df[grid_col].astype(str).str.match(r'^[A-Z0-9]{6,}$')].copy()
    col_map = {}
    col_map["name"]    = detect_column(df, ["Company / Account","Company","Account Name",
                                             "Name","restaurant_name","公司名称"])
    col_map["phone"]   = detect_column(df, ["Phone","phone_number","Mobile","电话"])
    col_map["street"]  = detect_column(df, ["Street","Formatted Restaurant Address",
                                             "restaurant_address","Address",
                                             "Block/Street Name","地址"])
    col_map["city"]    = detect_column(df, ["Area","City","restaurant_city","city",
                                             "Restaurant City","城市"])
    col_map["grid"]    = detect_column(df, ["GRID","grid","Grid"])
    col_map["lead_id"] = detect_column(df, ["Lead ID","lead_id","LeadID","ID"])
    col_map["url"]     = detect_column(df, ["GOOGLE URL","Google URL","google_url",
                                             "URL","Website"])
    col_map["lat"]     = detect_column(df, ["Coordinates (Latitude)","restaurant_lat",
                                             "Latitude","lat","纬度"])
    col_map["lng"]     = detect_column(df, ["Coordinates (Longitude)","restaurant_long",
                                             "Longitude","lng","经度"])
    col_map["zip"]     = detect_column(df, ["Zip/Postal Code","Restaurant PostalCode",
                                             "Zip","postal_code","PostalCode",
                                             "Postal Code","邮编"])
    return df, col_map


def load_crm(file_bytes: bytes, filename: str, market_cfg: dict):
    char_map = market_cfg["char_map"]
    df = _cached_read(file_bytes, filename)
    col_map = {}
    col_map["grid"]   = detect_column(df, ["GRID__c","GRID","Grid"])
    col_map["name"]   = detect_column(df, ["Account Name","Name","name"])
    col_map["phone"]  = detect_column(df, ["Phone","phone"])
    col_map["status"] = detect_column(df, ["Account_Status__c","Account Status",
                                            "AccountStatus"])
    col_map["reason"] = detect_column(df, ["Status_Reason__c","Status Reason",
                                            "StatusReason"])
    col_map["city"]   = detect_column(df, ["BillingCity","Restaurant City","City"])
    col_map["postal"] = detect_column(df, ["Restaurant PostalCode","PostalCode",
                                            "Postal Code","Zip/Postal Code",
                                            "BillingPostalCode","邮编"])
    col_map["street"] = detect_column(df, ["Formatted Restaurant Address",
                                            "BillingStreet","Street",
                                            "restaurant_address","Address",
                                            "Block/Street Name","地址"])
    # Heavy preprocessing — cached
    df = _cached_preprocess_crm(
        df,
        name_col   = col_map["name"],
        postal_col = col_map["postal"],
        addr_col   = col_map["street"],
        status_col = col_map["status"],
        char_map_tuple = tuple(sorted(char_map.items())),
    )
    return df, col_map


def load_apify(file_bytes: bytes, filename: str):
    df = _cached_read(file_bytes, filename)
    col_map = {}
    col_map["grid"]     = detect_column(df, ["GRID", "grid", "Grid"])
    col_map["title"]    = detect_column(df, ["title"])
    col_map["url"]      = detect_column(df, ["searchPageUrl", "search_page_url",
                                              "searchUrl", "input_url"])
    col_map["gm_url"]   = detect_column(df, ["url"])
    col_map["phone"]    = detect_column(df, ["phone"])
    col_map["website"]  = detect_column(df, ["website"])
    col_map["category"] = detect_column(df, ["categoryName", "category_name",
                                              "category", "categories/0"])
    col_map["perm"]     = detect_column(df, ["permanentlyClosed", "permanently_closed",
                                              "isClosed", "is_closed"])
    col_map["temp"]     = detect_column(df, ["temporarilyClosed", "temporarily_closed",
                                              "isTemporarilyClosed"])
    col_map["address"]  = detect_column(df, ["address"])
    col_map["lat"]      = detect_column(df, ["latitude","lat"])
    col_map["lng"]      = detect_column(df, ["longitude","lng"])
    if col_map["url"] and df[col_map["url"]].notna().sum() > 0:
        df["_url_norm"] = df[col_map["url"]].apply(norm_url)
    else:
        ss = detect_column(df, ["searchString"])
        if ss and df[ss].notna().sum() > 0:
            def _ex(s):
                if pd.isna(s): return ""
                s = str(s).strip()
                return norm_url(s.replace("Direct Detail URL:","").strip()
                                if "Direct Detail URL:" in s else s)
            df["_url_norm"] = df[ss].apply(_ex)
            col_map["url"] = ss
        elif col_map["gm_url"] and df[col_map["gm_url"]].notna().sum() > 0:
            df["_url_norm"] = df[col_map["gm_url"]].apply(norm_url)
            col_map["url"] = col_map["gm_url"]
    return df, col_map


# ═════════════════════════════════════════════════════════════════
# URL GENERATOR
# ═════════════════════════════════════════════════════════════════

def generate_google_urls(leads_df, col_map, market_cfg, mode="coords"):
    """
    mode="text"   → URL uses Company / Account + Street + Postal as query text
    mode="coords" → URL uses Coordinates (Latitude, Longitude) as query
    Falls back to the other mode if the preferred data is missing.
    """
    name_col   = col_map.get("name")
    street_col = col_map.get("street")
    zip_col    = col_map.get("zip")
    lat_col    = col_map.get("lat")
    lng_col    = col_map.get("lng")
    url_col    = col_map.get("url")
    suffix     = market_cfg.get("country_suffix","")

    urls, reused = [], 0
    for _, row in leads_df.iterrows():
        name   = str(row[name_col]).strip()   if name_col   and pd.notna(row.get(name_col))   else ""
        street = str(row[street_col]).strip() if street_col and pd.notna(row.get(street_col)) else ""
        postal = str(row[zip_col]).strip()    if zip_col    and pd.notna(row.get(zip_col))    else ""
        lat    = row.get(lat_col) if lat_col else None
        lng    = row.get(lng_col) if lng_col else None
        try:
            lat = float(lat) if lat is not None and str(lat) not in ("","nan") else None
            lng = float(lng) if lng is not None and str(lng) not in ("","nan") else None
        except (ValueError, TypeError):
            lat = lng = None

        # Reuse existing URL if present
        existing = str(row.get(url_col,"") or "") if url_col else ""
        if existing not in ("","nan") and "google.com/maps" in existing:
            urls.append(existing.strip()); reused += 1; continue

        url = ""

        if mode == "text":
            # Option 1: Company / Account + Street + Postal
            parts = [p for p in [name, street, postal] if p and p.lower() != "nan"]
            if parts:
                q = " ".join(parts)
                url = f"https://www.google.com/maps/search/?api=1&query={quote(q)}"
            elif lat and lng:
                # Fallback to coords if no text available
                url = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"

        else:
            # Option 2: Company / Account + Coordinates (Latitude, Longitude)
            if lat and lng:
                if name:
                    url = f"https://www.google.com/maps/search/?api=1&query={quote(f'{name},{lat},{lng}')}"
                else:
                    url = f"https://www.google.com/maps/search/?api=1&query={lat},{lng}"
            else:
                # Fallback to text if no coords available
                parts = [p for p in [name, street, postal] if p and p.lower() != "nan"]
                if parts:
                    q = " ".join(parts)
                    url = f"https://www.google.com/maps/search/?api=1&query={quote(q)}"
                elif name:
                    url = f"https://www.google.com/maps/search/?api=1&query={quote(f'{name} {suffix}')}"

        urls.append(url)
    return urls, reused


# ═════════════════════════════════════════════════════════════════
# GEO / ZONE
# ═════════════════════════════════════════════════════════════════

def geocode_address(street, postal_code, country_suffix, cache={}):
    key = f"{street}|{postal_code}|{country_suffix}"
    if key in cache: return cache[key]
    from urllib.request import Request
    street_c  = str(street).strip()      if street      and str(street).strip()      not in ("","nan") else ""
    postal_c  = str(postal_code).strip() if postal_code and str(postal_code).strip() not in ("","nan") else ""
    sg_postal = re.search(r'\b\d{6}\b', f"{street_c} {postal_c}")
    if sg_postal:
        token = get_onemap_token()
        if token:
            try:
                om = (f"https://www.onemap.gov.sg/api/common/elastic/search"
                      f"?searchVal={sg_postal.group(0)}&returnGeom=Y&getAddrDetails=N&pageNum=1")
                req = Request(om, headers={"Authorization": token,
                                           "User-Agent": "SalesOpsSuite/2.0"})
                with urlopen(req, timeout=5) as r:
                    res = json.loads(r.read().decode()).get("results",[])
                    if res:
                        lat, lng = float(res[0]["LATITUDE"]), float(res[0]["LONGITUDE"])
                        cache[key] = (lat, lng); return (lat, lng)
            except Exception: pass
    parts = [p for p in [street_c, postal_c, "Singapore"] if p.strip()]
    if not parts: cache[key] = (None, None); return (None, None)
    time.sleep(1.1)
    try:
        req = urlopen(
            f"https://photon.komoot.io/api/?q={urlencode({'q':', '.join(parts)})}&limit=1",
            timeout=6)
        feats = json.loads(req.read().decode()).get("features",[])
        if feats:
            c = feats[0]["geometry"]["coordinates"]
            cache[key] = (float(c[1]), float(c[0])); return cache[key]
    except Exception: pass
    cache[key] = (None, None); return (None, None)


def load_zones(file_bytes=None, filename=None, market_code=None):
    import os
    def _parse(df):
        zones, wc = [], next((c for c in df.columns if "wkt" in c.lower()), None)
        zn = next((c for c in df.columns if "zone_name" in c.lower()), None)
        if not wc: return zones
        for _, r in df.iterrows():
            w = str(r.get(wc,"")).strip()
            if not w or w.lower() == "nan": continue
            try: zones.append({"polygon": wkt_loads(w),
                                "zone_name": str(r.get(zn,"")) if zn else "",
                                "city_name": ""})
            except Exception: continue
        return zones
    if file_bytes:
        try:
            df = _cached_read(file_bytes, filename or "zones.csv")
            return _parse(df)
        except Exception: return []
    if market_code == "SG":
        p = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                         f"zones_{market_code}.json")
        if os.path.exists(p):
            try:
                raw = json.load(open(p, encoding="utf-8"))
                return [{"polygon": wkt_loads(z["wkt"]),
                         "zone_name": z.get("zone_name",""),
                         "city_name": z.get("city_name","")}
                        for z in raw if "wkt" in z]
            except Exception: return []
    return []


def point_in_zones(lat, lng, zones):
    if lat is None or lng is None: return None, None
    try:
        pt = Point(float(lng), float(lat))
        for z in zones:
            if z["polygon"].contains(pt):
                return z["zone_name"], z["city_name"]
    except Exception: pass
    return None, None


def check_delivery_zone(row, col_map, zones, suffix, geocode_enabled):
    if not zones: return "No Zone Data","","",""
    lat_c, lng_c = col_map.get("lat"), col_map.get("lng")
    str_c, zip_c = col_map.get("street"), col_map.get("zip")
    lat = row.get(lat_c) if lat_c else None
    lng = row.get(lng_c) if lng_c else None
    try:
        lat = float(lat) if lat is not None and not pd.isna(lat) else None
        lng = float(lng) if lng is not None and not pd.isna(lng) else None
    except (ValueError, TypeError):
        lat = lng = None
    if lat and lng:
        zn, zc = point_in_zones(lat, lng, zones)
        return (("Within Zone", zn, zc, "Coordinates")
                if zn else ("Outside Zone","","","Coordinates"))
    if geocode_enabled:
        street = str(row.get(str_c,"") or "") if str_c else ""
        postal = str(row.get(zip_c,"") or "") if zip_c else ""
        if street.strip() or postal.strip():
            lat, lng = geocode_address(street, postal, suffix)
            if lat:
                zn, zc = point_in_zones(lat, lng, zones)
                return (("Within Zone", zn, zc, "Geocoded")
                        if zn else ("Outside Zone","","","Geocoded"))
            return "Geocoding Failed","","","Geocoded"
    return "Outside Zone","","","No coordinates"


# ═════════════════════════════════════════════════════════════════
# CLASSIFY LEADS — MAIN ENGINE
# ═════════════════════════════════════════════════════════════════

def classify_leads(leads_df, col_map_leads, crm_df, col_map_crm,
                   apify_df, col_map_apify, market_cfg,
                   p2_threshold=0.50, p3_threshold=0.75,
                   exclusion_kw=None, zones=None, geocode_enabled=True,
                   progress_cb=None):

    char_map    = market_cfg["char_map"]
    prefix      = market_cfg["phone_prefix"]
    suffix      = market_cfg.get("country_suffix","")
    excl_kw     = exclusion_kw or _DEFAULT_EXCLUSION_KW

    # ── Build CRM indexes ─────────────────────────────────────────
    # Each item stored as (raw_name, lat_norm, pin_norm, row) so that
    # name normalisation is done ONCE here, not on every comparison.
    # crm_postal_all_dict gives O(1) pool lookup per lead postal.
    crm_postal_unit_dict = {}   # "postal|unit" -> [item, ...]
    crm_postal_all_dict  = {}   # postal -> [item, ...] (ALL at that postal)
    all_crm_items        = []   # flat list for zero-postal fallback

    if crm_df is not None:
        postal_c = col_map_crm.get("postal")
        addr_c   = col_map_crm.get("street")
        name_c_  = col_map_crm.get("name")

        for _, r in crm_df.iterrows():
            raw_n         = str(r.get(name_c_,"") or "") if name_c_ else ""
            lat_n, pin_n  = norm_name_sg(strip_venue_generic(raw_n), char_map)
            item          = (raw_n, lat_n, pin_n, r)
            all_crm_items.append(item)

            postal_raw = _norm_postal_input(r.get(postal_c,"") if postal_c else "")
            addr_raw   = str(r.get(addr_c,  "") or "") if addr_c   else ""
            crm_postal = extract_sg_postal(postal_raw) or extract_sg_postal(addr_raw)
            crm_unit   = extract_sg_unit(addr_raw)

            if crm_postal:
                crm_postal_all_dict.setdefault(crm_postal, []).append(item)
                if crm_unit:
                    crm_postal_unit_dict.setdefault(
                        f"{crm_postal}|{crm_unit}", []).append(item)

    # ── Build Apify GRID index ─────────────────────────────────────
    apify_dict = {}
    if apify_df is not None:
        gc = detect_column(apify_df, ["GRID", "grid", "Grid"])
        if gc:
            for _, r in apify_df.iterrows():
                g = str(r.get(gc,"") or "").strip()
                if g and g not in apify_dict:
                    apify_dict[g] = r

    # ── Column aliases ─────────────────────────────────────────────
    name_col_l   = col_map_leads.get("name")
    phone_col_l  = col_map_leads.get("phone")
    street_col_l = col_map_leads.get("street")
    lat_col_l    = col_map_leads.get("lat")
    lng_col_l    = col_map_leads.get("lng")
    grid_col_l   = col_map_leads.get("grid")
    lead_id_col  = col_map_leads.get("lead_id")
    zip_col_l    = col_map_leads.get("zip")
    reason_col_c = col_map_crm.get("reason") if col_map_crm else None

    results = []
    _crm_name_col   = col_map_crm.get("name","")   if col_map_crm else ""
    _crm_grid_col   = col_map_crm.get("grid","")   if col_map_crm else ""
    _crm_status_col = col_map_crm.get("status","") if col_map_crm else ""
    _crm_street_col = col_map_crm.get("street","") if col_map_crm else ""
    _crm_postal_col = col_map_crm.get("postal","") if col_map_crm else ""
    _n_total        = len(leads_df)

    # ── Scoring helper (defined ONCE — not recreated per lead) ────
    def _score(lead_lat_n, lead_pin_n, item):
        """Score a pre-normalised CRM item against a lead.
        Returns (score, crm_raw, crm_row).
        Uses pre-computed lat_n / pin_n — avoids re-normalising on
        every comparison (critical for 100k+ CRM accounts)."""
        raw_n, lat_n, pin_n, r = item
        ns = 0.0
        if lead_lat_n and lat_n:
            ns = max(fuzz.token_sort_ratio(lead_lat_n, lat_n),
                     fuzz.token_set_ratio(lead_lat_n,  lat_n)) / 100.0
        if lead_pin_n or pin_n:
            lp, cp = lead_pin_n or "", pin_n or ""
            if lp and cp:
                ps = max(fuzz.token_sort_ratio(lp, cp),
                         fuzz.token_set_ratio(lp,  cp)) / 100.0
                return round(max(ns, ps), 3), raw_n, r
        return round(ns, 3), raw_n, r

    def _pool(postal):
        """O(1) pool lookup — returns all CRM items at this postal."""
        return crm_postal_all_dict.get(postal, [])

    for _lead_i, (_, row) in enumerate(leads_df.iterrows()):

        # ── Extract lead fields ────────────────────────────────────
        lead_name_raw = row.get(name_col_l,"")   if name_col_l   else ""
        lead_street   = str(row.get(street_col_l,"") or "") if street_col_l else ""
        lead_zip      = str(row.get(zip_col_l,  "") or "") if zip_col_l    else ""

        lead_lat = row.get(lat_col_l) if lat_col_l else None
        lead_lng = row.get(lng_col_l) if lng_col_l else None
        try:
            lead_lat = float(lead_lat) if lead_lat is not None and not pd.isna(lead_lat) else None
            lead_lng = float(lead_lng) if lead_lng is not None and not pd.isna(lead_lng) else None
        except (ValueError, TypeError):
            lead_lat = lead_lng = None

        lead_postal_raw = _norm_postal_input(row.get(zip_col_l,"") if zip_col_l else "")
        lead_postal = (extract_sg_postal(lead_postal_raw)
                       or extract_sg_postal(lead_street))
        lead_unit   = extract_sg_unit(lead_street)
        lead_lat_n, lead_pin_n = norm_name_sg(strip_venue_generic(lead_name_raw), char_map)
        lead_blank  = is_blank_name(lead_name_raw)
        lead_grid   = str(row.get(grid_col_l,"") or "").strip() if grid_col_l else ""

        # ── Initialise ─────────────────────────────────────────────
        crm_match          = None
        match_method       = ""
        match_score        = 0.0
        label              = ""
        dup_grid = dup_name = dup_crm_status = dup_reason = dup_method = ""
        dup_address = dup_postal = ""
        prev_occupant_grid = ""
        prev_occupant_name = ""

        # ── Pre-flight: exclude Mix & Match immediately ────────────
        if not lead_blank and "mix & match" in str(lead_name_raw).lower():
            label        = "Wrong Target Group"
            match_method = "Lead name contains 'Mix & Match' — excluded"

        # ══════════════════════════════════════════════════════════
        # CRM DEDUP — Postal + Name (unit used as filter)
        #
        # Logic:
        #   1. If lead has a unit AND CRM has an exact unit match →
        #      score THAT record. If name < p2, it's a new tenant.
        #   2. Otherwise → score all CRM accounts at same postal,
        #      take best name match.
        #   score ≥ p3_threshold (0.75) → P4 Duplicate
        #   score  p2–p3 (0.50–0.74)  → P3 Potential Match
        #   score < p2_threshold (0.50) → no CRM match → Apify
        # ══════════════════════════════════════════════════════════
        if not label and lead_postal and not lead_blank:

            if lead_unit:
                unit_matches = crm_postal_unit_dict.get(f"{lead_postal}|{lead_unit}", [])

                if unit_matches:
                    best_sc, best_cand_row, best_raw = 0.0, None, ""
                    for item in unit_matches:
                        sc, crm_raw, cand_row = _score(lead_lat_n, lead_pin_n, item)
                        if sc > best_sc:
                            best_sc, best_cand_row, best_raw = sc, cand_row, crm_raw

                    if best_sc >= p2_threshold:
                        crm_match    = best_cand_row
                        match_score  = best_sc
                        match_method = (f"Postal+Unit+Name [{lead_postal} #{lead_unit}] "
                                        f"score={best_sc:.2f}")
                    else:
                        prev_occupant_name = best_raw
                        prev_occupant_grid = str(
                            best_cand_row.get(_crm_grid_col,"") or "") \
                            if best_cand_row is not None else ""
                        match_method = (f"New business at known address "
                                        f"[{lead_postal} #{lead_unit}] "
                                        f"prev='{best_raw}' score={best_sc:.2f}")
                else:
                    # No CRM record at lead unit → search all at same postal
                    # 0.90 minimum since unit is unconfirmed
                    no_unit_min = max(p2_threshold, 0.90)
                    pool = _pool(lead_postal)
                    best_sc, best_cand_row, best_raw = 0.0, None, ""
                    for item in pool:
                        sc, crm_raw, cand_row = _score(lead_lat_n, lead_pin_n, item)
                        if sc > best_sc:
                            best_sc, best_cand_row, best_raw = sc, cand_row, crm_raw
                    if best_cand_row is not None and best_sc >= no_unit_min:
                        crm_match    = best_cand_row
                        match_score  = best_sc
                        match_method = (f"Postal+Name [{lead_postal}] "
                                        f"score={best_sc:.2f}")
            else:
                # Lead has no unit → postal + name only, 0.90 minimum
                no_unit_min = max(p2_threshold, 0.90)
                pool = _pool(lead_postal)
                best_sc, best_cand_row, best_raw = 0.0, None, ""
                for item in pool:
                    sc, crm_raw, cand_row = _score(lead_lat_n, lead_pin_n, item)
                    if sc > best_sc:
                        best_sc, best_cand_row, best_raw = sc, cand_row, crm_raw
                if best_cand_row is not None and best_sc >= no_unit_min:
                    crm_match    = best_cand_row
                    match_score  = best_sc
                    match_method = (f"Postal+Name [{lead_postal}] "
                                    f"score={best_sc:.2f} (no unit)")

        # ── Assign CRM-based label ─────────────────────────────────
        if crm_match is not None:
            label          = ("P4 — Duplicate"      if match_score >= p3_threshold
                              else "P3 — Potential Match")
            dup_grid       = str(crm_match.get(_crm_grid_col,"") or "")
            dup_name       = str(crm_match.get(_crm_name_col,"") or "")
            dup_crm_status = str(crm_match.get(_crm_status_col,"") or "")
            dup_reason     = str(crm_match.get(reason_col_c,"") or "") if reason_col_c else ""
            dup_method     = match_method
            dup_address    = str(crm_match.get(_crm_street_col,"") or "")
            dup_postal     = str(crm_match.get(_crm_postal_col,"") or "")

        # ══════════════════════════════════════════════════════════
        # ZERO-POSTAL FALLBACK — Name-only scan (full CRM)
        #
        # Triggered when the lead has no valid SG postal code
        # (000000 / 00000 / blank) — meaning the address data is
        # unreliable and the postal cascade produced no match.
        #
        # Rules:
        #   • Score ≥ 0.90 (hard minimum, not adjustable from sidebar)
        #   • Geographic conflict check: if CRM has an Area: line and
        #     that area conflicts with a known area in the lead text,
        #     reject the candidate.
        #   • Result is always P3 (no postal confirmation available).
        # ══════════════════════════════════════════════════════════
        if crm_match is None and not label and not lead_blank and not lead_postal:

            def _extract_crm_area(addr_raw: str) -> str:
                m = re.search(r'Area:\s*(.+)', str(addr_raw), re.IGNORECASE)
                return m.group(1).strip().lower() if m else ""

            def _geo_conflict(lead_text: str, crm_area: str) -> bool:
                if not crm_area: return False
                lead_low = lead_text.lower()
                if crm_area in lead_low: return False
                return any(area in lead_low for area in SG_AREAS if area != crm_area)

            lead_text_for_geo   = f"{lead_name_raw} {lead_street}".lower()
            NAME_ONLY_THRESHOLD = 0.90
            best_sc, best_cand  = 0.0, None

            for item in all_crm_items:
                sc, _, cand_row = _score(lead_lat_n, lead_pin_n, item)
                if sc < NAME_ONLY_THRESHOLD or sc <= best_sc:
                    continue
                addr_raw_c = str(cand_row.get(_crm_street_col,"") or "")
                if _geo_conflict(lead_text_for_geo, _extract_crm_area(addr_raw_c)):
                    continue
                best_sc, best_cand = sc, cand_row

            if best_cand is not None:
                crm_match      = best_cand
                match_score    = best_sc
                match_method   = f"Name-only (zero postal) score={best_sc:.2f}"
                label          = "P3 — Potential Match"
                dup_grid       = str(crm_match.get(_crm_grid_col,"") or "")
                dup_name       = str(crm_match.get(_crm_name_col,"") or "")
                dup_crm_status = str(crm_match.get(_crm_status_col,"") or "")
                dup_reason     = str(crm_match.get(reason_col_c,"") or "") if reason_col_c else ""
                dup_method     = match_method
                dup_address    = str(crm_match.get(_crm_street_col,"") or "")
                dup_postal     = str(crm_match.get(_crm_postal_col,"") or "")

        # ══════════════════════════════════════════════════════════
        # APIFY VALIDATION — GRID exact lookup
        # ══════════════════════════════════════════════════════════
        apy = apify_dict.get(lead_grid) if lead_grid else None
        gm_title = gm_cat = gm_biz_status = gm_phone = gm_website = gm_url = ""
        match_conf = 0.0; match_reason = ""

        if apy is not None:
            tc  = col_map_apify.get("title");    gm_title   = str(apy[tc] or "")  if tc  and pd.notna(apy.get(tc))  else ""
            cc  = col_map_apify.get("category"); gm_cat     = str(apy[cc] or "")  if cc  and pd.notna(apy.get(cc))  else ""
            pc  = col_map_apify.get("phone");    gm_phone   = to_e164(apy.get(pc,""), prefix) if pc else ""
            wc  = col_map_apify.get("website");  gm_website = str(apy[wc] or "")  if wc  and pd.notna(apy.get(wc))  else ""
            uc  = col_map_apify.get("gm_url");   gm_url     = str(apy[uc] or "")  if uc  and pd.notna(apy.get(uc))  else ""
            prc = col_map_apify.get("perm");     perm = str(apy.get(prc,"")).lower() == "true" if prc else False
            tmc = col_map_apify.get("temp");     temp = str(apy.get(tmc,"")).lower() == "true" if tmc else False
            gm_biz_status = ("Permanently Closed" if perm else
                             "Temporarily Closed" if temp else "Open")

            gm_phone_norm = norm_phone(apy.get(pc,""), prefix) if pc else ""
            name_sc       = name_confidence(lead_name_raw, gm_title, char_map)
            addr_col_a    = col_map_apify.get("address","")
            addr_hit      = address_match(lead_street, apy.get(addr_col_a,""), char_map)

            match_conf = round((min(name_sc, 1.0) * 0.6)
                               + (0.4 if addr_hit else 0.0), 3)
            # GRID guarantees we have the right Apify row — confirmed by definition.
            confirmed  = True

            reasons = []
            reasons.append(f"Name {name_sc:.2f}" + (" ✓" if name_sc >= 0.5 else " ✗"))
            if addr_hit: reasons.append("Address ✓")
            reasons.append("GRID match ✓")
            match_reason = " | ".join(reasons)

            # Apify sets label only when CRM cascade found no match.
            # Two safety checks prevent wrong Apify results from mislabelling:
            #   1. If Apify says closed BUT address doesn't match the lead →
            #      likely a different branch/location → P2
            #   2. If category is non-food AND name barely matches the lead →
            #      Apify returned the wrong business entirely → P2
            if not label:
                wrong_location   = (perm or temp) and not addr_hit
                wrong_business   = (not is_food_delivery_eligible(gm_cat, excl_kw)
                                    and name_sc < 0.2)
                if wrong_location or wrong_business:
                    label = "P2 — Please Check"
                elif not gm_cat:                                      label = "P2 — Please Check"
                elif perm or temp:                                    label = "Business Closed"
                elif not is_food_delivery_eligible(gm_cat, excl_kw): label = "Wrong Target Group"
                else:                                                 label = "P1 — New"
        else:
            gm_biz_status = "Not Found on Google"
            match_reason  = "No Apify result"
            if not label: label = "P2 — Please Check"

        # ── Zone check ─────────────────────────────────────────────
        zone_status = zone_name = zone_city = zone_method = ""
        if zones:
            zone_status, zone_name, zone_city, zone_method = check_delivery_zone(
                row, col_map_leads, zones, suffix, geocode_enabled)

        results.append({
            "GRID":                   row.get(grid_col_l,"")   if grid_col_l  else "",
            "Lead ID":                row.get(lead_id_col,"")  if lead_id_col else "",
            "Company / Account":      lead_name_raw,
            "City":                   "SINGAPORE",
            "Street":                 lead_street,
            "Phone":                  to_e164(row.get(phone_col_l,""), prefix) if phone_col_l else "",
            "GM Title":               gm_title,
            "GM Category":            gm_cat,
            "GM Business Status":     gm_biz_status,
            "GM Phone":               gm_phone,
            "GM Website":             gm_website,
            "GM URL":                 gm_url,
            "Match Confidence":       match_conf,
            "Match Reason":           match_reason,
            "Label":                  label,
            "Match Score (CRM)":      round(match_score, 3),
            "Duplicate GRID":         dup_grid,
            "Duplicate CRM Name":     dup_name,
            "Duplicate CRM Address":  dup_address,
            "Duplicate CRM Postal":   dup_postal,
            "CRM Account Status":     dup_crm_status,
            "CRM Status Reason":      dup_reason,
            "Duplicate Match Method": dup_method,
            "Previous Occupant GRID": prev_occupant_grid,
            "Previous Occupant Name": prev_occupant_name,
            "Delivery Zone Status":   zone_status,
            "Zone Name":              zone_name,
            "Zone City":              zone_city,
            "Zone Method":            zone_method,
        })

        if progress_cb and (_lead_i % 10 == 0 or _lead_i == _n_total - 1):
            progress_cb(_lead_i + 1, _n_total)

    return pd.DataFrame(results)


# ═════════════════════════════════════════════════════════════════
# SF ACCOUNT AUDIT  (finds duplicates WITHIN Salesforce)
# ═════════════════════════════════════════════════════════════════

def find_sf_duplicates(df, name_col, addr_col, status_col, grid_col, id_col, threshold):
    """
    Find suspected duplicate account pairs within the SF master.
    Trigger: same 6-digit postal code + same extracted unit number
             + name similarity ≥ threshold.
    """
    pairs = []
    for postal, pg in df.groupby("_postal_fixed"):
        with_unit = pg[pg["_unit_extracted"] != ""]
        if with_unit.empty: continue
        for unit, ug in with_unit.groupby("_unit_extracted"):
            if len(ug) < 2: continue
            rows = ug.reset_index(drop=True)
            for i, j in combinations(range(len(rows)), 2):
                a, b   = rows.iloc[i], rows.iloc[j]
                score  = fuzz.token_sort_ratio(a["_name_latin"], b["_name_latin"])
                if score < threshold: continue
                risk, action = get_risk_and_action(a[status_col], b[status_col])
                pairs.append({
                    "ACCT_A_SF_ID":       str(a[id_col]),
                    "ACCT_A_NAME":        str(a[name_col]),
                    "ACCT_A_STATUS":      str(a[status_col]),
                    "ACCT_A_ADDRESS":     str(a[addr_col]) if addr_col else "",
                    "ACCT_A_POSTAL":      str(a["_postal_fixed"]),
                    "ACCT_A_GRID":        str(a[grid_col]) if grid_col else "",
                    "ACCT_B_SF_ID":       str(b[id_col]),
                    "ACCT_B_NAME":        str(b[name_col]),
                    "ACCT_B_STATUS":      str(b[status_col]),
                    "ACCT_B_ADDRESS":     str(b[addr_col]) if addr_col else "",
                    "ACCT_B_POSTAL":      str(b["_postal_fixed"]),
                    "ACCT_B_GRID":        str(b[grid_col]) if grid_col else "",
                    "SHARED_UNIT":        unit,
                    "NAME_SCORE":         f"{score}%",
                    "RISK_LEVEL":         risk,
                    "RECOMMENDED_ACTION": action,
                })
    return pairs


# ═════════════════════════════════════════════════════════════════
# EXCEL REPORT BUILDER
# ═════════════════════════════════════════════════════════════════

# ─────────────────────────────────────────────────────────────────────
# CRM CHECK — standalone duplicate check (no GRID, no Apify needed)
# ─────────────────────────────────────────────────────────────────────

def crm_check_classify(rest_df, rest_cols, crm_df, col_map_crm,
                        char_map, p2_threshold, p3_threshold):
    """
    CRM-only duplicate check for raw restaurant lists.
    Input: name + street + postal only. No GRID or Apify needed.
    Labels: P4 — Duplicate / P3 — Potential Match / Unverified
    """
    # ── Build CRM indexes (pre-compute names once) ────────────────
    crm_postal_unit_dict = {}   # "postal|unit" -> [item, ...]
    crm_postal_all_dict  = {}   # postal -> [item, ...]
    all_crm_items        = []

    crm_name_c = col_map_crm.get("name")   if col_map_crm else None
    crm_grid_c = col_map_crm.get("grid")   if col_map_crm else None
    crm_stat_c = col_map_crm.get("status") if col_map_crm else None
    crm_reas_c = col_map_crm.get("reason") if col_map_crm else None
    crm_addr_c = col_map_crm.get("street") if col_map_crm else None

    if crm_df is not None:
        postal_c = col_map_crm.get("postal") if col_map_crm else None
        addr_c   = crm_addr_c
        for _, r in crm_df.iterrows():
            raw_n        = str(r.get(crm_name_c,"") or "") if crm_name_c else ""
            lat_n, pin_n = norm_name_sg(strip_venue_generic(raw_n), char_map)
            item         = (raw_n, lat_n, pin_n, r)
            all_crm_items.append(item)
            postal_raw = _norm_postal_input(r.get(postal_c,"") if postal_c else "")
            addr_raw   = str(r.get(addr_c,  "") or "") if addr_c   else ""
            crm_postal = extract_sg_postal(postal_raw) or extract_sg_postal(addr_raw)
            crm_unit   = extract_sg_unit(addr_raw)
            if crm_postal:
                crm_postal_all_dict.setdefault(crm_postal, []).append(item)
                if crm_unit:
                    crm_postal_unit_dict.setdefault(
                        f"{crm_postal}|{crm_unit}", []).append(item)

    name_col   = rest_cols.get("name")
    street_col = rest_cols.get("street")
    postal_col = rest_cols.get("postal")
    grid_col   = rest_cols.get("grid")

    # ── Helpers (defined once) ────────────────────────────────────
    def _sc(lead_lat, lead_pin, item):
        raw_n, lat_n, pin_n, r = item
        ns = 0.0
        if lead_lat and lat_n:
            ns = max(fuzz.token_sort_ratio(lead_lat, lat_n),
                     fuzz.token_set_ratio(lead_lat,  lat_n)) / 100.0
        if lead_pin or pin_n:
            lp, cp = lead_pin or "", pin_n or ""
            if lp and cp:
                ps = max(fuzz.token_sort_ratio(lp, cp),
                         fuzz.token_set_ratio(lp,  cp)) / 100.0
                return round(max(ns, ps), 3), raw_n, r
        return round(ns, 3), raw_n, r

    def _crm_n(r): return str(r.get(crm_name_c,"") or "") if crm_name_c else ""

    def _extract_area(addr_raw):
        m = re.search(r'Area:\s*(.+)', str(addr_raw), re.IGNORECASE)
        return m.group(1).strip().lower() if m else ""

    def _geo_conflict(lead_text, crm_area):
        if not crm_area: return False
        lead_low = lead_text.lower()
        if crm_area in lead_low: return False
        return any(area in lead_low for area in SG_AREAS if area != crm_area)

    results = []

    for _, row in rest_df.iterrows():
        name_raw = str(row.get(name_col,"") or "") if name_col else ""
        street   = str(row.get(street_col,"") or "") if street_col else ""
        postal_s = str(row.get(postal_col,"") or "") if postal_col else ""
        grid_val = str(row.get(grid_col,"") or "") if grid_col else ""
        postal  = extract_sg_postal(_norm_postal_input(postal_s)) or extract_sg_postal(street)
        unit    = extract_sg_unit(street)
        blank   = is_blank_name(name_raw)
        lat_n, pin_n = norm_name_sg(strip_venue_generic(name_raw), char_map)

        crm_match = None; match_score = 0.0; match_method = ""

        if postal and not blank:
            if unit:
                unit_matches = crm_postal_unit_dict.get(f"{postal}|{unit}", [])
                pool = unit_matches if unit_matches else crm_postal_all_dict.get(postal, [])
                tag  = f"Postal+Unit+Name [{postal} #{unit}]" if unit_matches \
                       else f"Postal+Name [{postal}]"
                # No-unit fallback uses stricter threshold
                min_sc = p2_threshold if unit_matches else max(p2_threshold, 0.90)
            else:
                pool   = crm_postal_all_dict.get(postal, [])
                tag    = f"Postal+Name [{postal}] (no unit)"
                min_sc = max(p2_threshold, 0.90)  # 0.90 minimum for no-unit

            best_sc, best_cand, best_raw = 0.0, None, ""
            for item in pool:
                sc, raw_n, cand_row = _sc(lat_n, pin_n, item)
                if sc > best_sc:
                    best_sc, best_cand, best_raw = sc, cand_row, raw_n
            if best_cand is not None and best_sc >= min_sc:
                crm_match = best_cand; match_score = best_sc
                match_method = f"{tag} score={best_sc:.2f}"
            elif best_cand is not None:
                match_method = f"No name match at {tag} best='{best_raw}' score={best_sc:.2f}"
            else:
                match_method = f"No CRM accounts at postal {postal}"

        elif blank:
            match_method = "Blank restaurant name — skipped"

        else:
            # Zero/no postal → name-only scan at ≥0.90
            lead_geo = f"{name_raw} {street}".lower()
            best_sc, best_cand = 0.0, None
            for item in all_crm_items:
                sc, _, cand_row = _sc(lat_n, pin_n, item)
                if sc < 0.90 or sc <= best_sc: continue
                addr_raw_c = str(cand_row.get(crm_addr_c,"") or "") if crm_addr_c else ""
                if _geo_conflict(lead_geo, _extract_area(addr_raw_c)): continue
                best_sc, best_cand = sc, cand_row
            if best_cand is not None:
                crm_match = best_cand; match_score = best_sc
                match_method = f"Name-only (zero postal) score={best_sc:.2f}"
            else:
                match_method = "No postal — name-only scan: no match at ≥90%"

        # ── Label ─────────────────────────────────────────────────
        if crm_match is not None:
            label      = ("P4 — Duplicate" if match_score >= p3_threshold
                          else "P3 — Potential Match")
            dup_grid   = str(crm_match.get(crm_grid_c,"") or "") if crm_grid_c else ""
            dup_name   = _crm_n(crm_match)
            dup_status = str(crm_match.get(crm_stat_c,"") or "") if crm_stat_c else ""
            dup_reas   = str(crm_match.get(crm_reas_c,"") or "") if crm_reas_c else ""
            dup_addr   = str(crm_match.get(crm_addr_c,"") or "") if crm_addr_c else ""
            dup_post   = str(crm_match.get(col_map_crm.get("postal",""),"") or "") \
                         if col_map_crm else ""
        else:
            label = "Unverified"
            dup_grid = dup_name = dup_status = dup_reas = dup_addr = dup_post = ""

        results.append({
            "GRID":               grid_val,
            "Company / Account":  name_raw,
            "Street":             street,
            "Zip/Postal Code":    postal_s,
            "Label":              label,
            "Match Score":        match_score if crm_match is not None else "",
            "Duplicate CRM Name": dup_name,
            "Duplicate GRID":     dup_grid,
            "Duplicate CRM Address": dup_addr,
            "Duplicate CRM Postal":  dup_post,
            "CRM Account Status": dup_status,
            "CRM Status Reason":  dup_reas,
            "Match Reason":       match_method,
        })

    return pd.DataFrame(results)


def build_crm_check_excel(df: pd.DataFrame) -> bytes:
    from io import BytesIO

    FILLS_C = {
        "P4 — Duplicate":       PatternFill("solid", start_color="FFC7CE"),
        "P3 — Potential Match": PatternFill("solid", start_color="FFF2CC"),
        "Unverified":           PatternFill("solid", start_color="DBEAFE"),
    }
    ALT_C = {
        "P4 — Duplicate":       PatternFill("solid", start_color="FFE0E0"),
        "P3 — Potential Match": PatternFill("solid", start_color="FFFAE0"),
        "Unverified":           PatternFill("solid", start_color="EFF6FF"),
    }
    FC_C = {
        "P4 — Duplicate":       "9C0006",
        "P3 — Potential Match": "7D5A00",
        "Unverified":           "1E40AF",
    }
    HDR_FILL = PatternFill("solid", start_color="1F4E79")

    def _thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    heads  = ["GRID","Company / Account","Street","Zip/Postal Code","Label",
              "Match Score","Duplicate CRM Name","Duplicate GRID",
              "Duplicate CRM Address","Duplicate CRM Postal",
              "CRM Account Status","CRM Status Reason","Match Reason"]
    widths = [10,32,36,14,22,12,30,14,36,14,16,18,40]
    widths = [32,36,14,22,12,30,14,18,18,45]

    def _build_sheet(ws, title, data):
        ws["A1"] = title
        ws["A1"].font = Font(name="Poppins", bold=True, size=12, color="1F4E79")
        ws.merge_cells(f"A1:{get_column_letter(len(heads))}1")
        for ci, h in enumerate(heads, 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font      = Font(name="Poppins", bold=True, color="FFFFFF", size=9)
            c.fill      = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _thin()
        ws.row_dimensions[3].height = 30
        for ri, (_, row) in enumerate(data.iterrows(), 4):
            lbl  = row["Label"]
            fill = FILLS_C.get(lbl, PatternFill()) if ri % 2 == 0 else ALT_C.get(lbl, PatternFill())
            fc   = FC_C.get(lbl, "000000")
            ws.row_dimensions[ri].height = 20
            for ci, key in enumerate(heads, 1):
                val = row.get(key,""); val = "" if pd.isna(val) else val
                c   = ws.cell(row=ri, column=ci, value=val)
                c.border = _thin()
                if key == "Label":
                    c.font = Font(name="Poppins", size=8, bold=True, color=fc)
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif key == "Match Score":
                    try: c.value = float(val); c.number_format = "0.00"
                    except Exception: pass
                    c.font = Font(name="Poppins", size=8, bold=True)
                    c.fill = fill
                    c.alignment = Alignment(horizontal="center", vertical="center")
                elif key == "Match Reason":
                    c.font = Font(name="Poppins", size=8, color="595959")
                    c.fill = fill
                    c.alignment = Alignment(vertical="center", wrap_text=True)
                else:
                    c.font = Font(name="Poppins", size=8)
                    c.fill = fill
                    c.alignment = Alignment(vertical="center")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"
        if len(data) > 0:
            ws.auto_filter.ref = f"A3:{get_column_letter(len(heads))}{3 + len(data)}"

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "All Results"
    _build_sheet(ws1, "CRM Duplicate Check — All Results", df)

    ws2 = wb.create_sheet("✅ Unverified — Create")
    _build_sheet(ws2, "Unverified — No CRM Match (Safe to Create)",
                 df[df["Label"] == "Unverified"].reset_index(drop=True))

    ws3 = wb.create_sheet("🟡 P3 — Review First")
    _build_sheet(ws3, "P3 — Potential Match (Review Before Creating)",
                 df[df["Label"] == "P3 — Potential Match"].reset_index(drop=True))

    ws4 = wb.create_sheet("🔴 P4 — Duplicates")
    _build_sheet(ws4, "P4 — Duplicate (Already in Salesforce — Skip)",
                 df[df["Label"] == "P4 — Duplicate"].reset_index(drop=True))

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def build_excel(df: pd.DataFrame, market_name: str):
    # ── Color palette ──────────────────────────────────────────────
    FILLS = {
        "P1 — New":            PatternFill("solid", start_color="C6EFCE"),
        "P2 — Please Check":   PatternFill("solid", start_color="ECECEC"),
        "P3 — Potential Match":PatternFill("solid", start_color="FFF2CC"),
        "P4 — Duplicate":      PatternFill("solid", start_color="FFC7CE"),
        "Business Closed":     PatternFill("solid", start_color="FFEB9C"),
        "Wrong Target Group":  PatternFill("solid", start_color="FFDCA8"),
    }
    ALT = {k: PatternFill("solid", start_color=
           {"P1 — New":"EBF7EB","P2 — Please Check":"F7F7F7",
            "P3 — Potential Match":"FFFAE0","P4 — Duplicate":"FFE0E0",
            "Business Closed":"FFF7D1","Wrong Target Group":"FFF0DC"}.get(k,"FFFFFF"))
           for k in FILLS}
    FONT_C = {
        "P1 — New":"276221","P2 — Please Check":"595959",
        "P3 — Potential Match":"7D5A00","P4 — Duplicate":"9C0006",
        "Business Closed":"7D4E00","Wrong Target Group":"833C00",
    }
    HDR_FILL = PatternFill("solid", start_color="1F4E79")
    SEC_FILL = PatternFill("solid", start_color="2E75B6")

    def _thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)
    def _hdr(ws, r, c, v):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Poppins", bold=True, color="FFFFFF", size=9)
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _thin()
    def _sec(ws, r, c, text, span=3):
        ws.merge_cells(start_row=r, start_column=c,
                       end_row=r, end_column=c + span - 1)
        cell = ws.cell(row=r, column=c, value=f"  {text}")
        cell.font      = Font(name="Poppins", bold=True, color="FFFFFF", size=9)
        cell.fill      = SEC_FILL
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[r].height = 22
    def _dc(ws, r, c, v, fill=None, bold=False, fmt=None,
             align="center", color="000000"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font      = Font(name="Poppins", size=8, bold=bold, color=color)
        cell.fill      = fill if fill else PatternFill()
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = _thin()
        if fmt: cell.number_format = fmt

    def _fill(lbl, ri):
        base = FILLS.get(lbl, PatternFill())
        alt  = ALT.get(lbl, PatternFill())
        return base if ri % 2 == 0 else alt

    def _conf_fill(score):
        try: score = float(score)
        except Exception: return PatternFill()
        if score >= 0.8: return PatternFill("solid", start_color="C6EFCE")
        if score >= 0.6: return PatternFill("solid", start_color="FFEB9C")
        return PatternFill("solid", start_color="FFC7CE")

    labels_order = [
        "P1 — New", "P3 — Potential Match", "P4 — Duplicate",
        "Business Closed", "Wrong Target Group", "P2 — Please Check",
    ]
    col_headers = [
        # ── A–F: Agent workflow columns (blank on export, filled by agents) ──
        "Agent","Due Date","Convert/Lost","Invalid Reason",
        "Comments/Duplicate GRID","Feedback",
        # ── G onwards: classifier output ─────────────────────────────────────
        "GRID","Lead ID","Company / Account","City","Street","Phone",   # G–L
        "GM Title","GM Category","GM Business Status",                  # M–O
        "GM Phone","GM Website","GM URL",                               # P–R
        "Match Confidence","Match Reason","Label","Match Score (CRM)",  # S–V
        "Duplicate GRID","Duplicate CRM Name",                          # W–X
        "Duplicate CRM Address","Duplicate CRM Postal",                 # Y–Z (new)
        "CRM Account Status","CRM Status Reason","Duplicate Match Method", # AA–AC
        "Previous Occupant GRID","Previous Occupant Name",              # AD–AE
        "Delivery Zone Status","Zone Name","Zone City","Zone Method",    # AF–AI
    ]
    col_w = [
        10,12,14,18,22,20,              # Agent cols A–F
        10,18,32,12,36,16,              # G–L (GRID … Phone)
        28,24,18,16,28,48,              # M–R (GM Title … GM URL)
        14,38,24,14,14,30,              # S–V (Match Conf … Match Score)
        16,18,36,14,                    # W–Z (Dup GRID, Dup Name, Dup Addr, Dup Postal)
        16,18,40,14,30,18,18,14,14,    # AA–AI (CRM Status … Zone Method)
    ]

    DATA_S = 5
    DATA_E = DATA_S + len(df) - 1
    LBL_R  = f"'Classified Leads'!U{DATA_S}:U{DATA_E}"   # col 21 = U (unchanged)
    CAT_R  = f"'Classified Leads'!N{DATA_S}:N{DATA_E}"   # col 14 = N (unchanged)
    CRMS_R = f"'Classified Leads'!AA{DATA_S}:AA{DATA_E}" # col 27 = AA (was Y, +2)
    METH_R = f"'Classified Leads'!AC{DATA_S}:AC{DATA_E}" # col 29 = AC (was AA, +2)
    ZONE_R = f"'Classified Leads'!AF{DATA_S}:AF{DATA_E}" # col 32 = AF (was AD, +2)
    has_z  = df["Delivery Zone Status"].astype(str).str.strip().ne("").any()

    wb  = Workbook()

    # ── Sheet 1: Classified Leads ─────────────────────────────────
    ws1 = wb.active; ws1.title = "Classified Leads"
    counts = df["Label"].value_counts()
    ws1["A1"] = f"Lead Classification Report  |  {market_name}"
    ws1["A1"].font = Font(name="Poppins", bold=True, size=14, color="1F4E79")
    ws1.merge_cells("A1:AG1")
    ws1["A2"] = "Total: {:,}   |   {}".format(
        len(df), "   ".join(f"{l}: {counts.get(l,0):,}" for l in labels_order))
    ws1["A2"].font = Font(name="Poppins", italic=True, size=8, color="595959")
    ws1.merge_cells("A2:AG2")
    for ci, h in enumerate(col_headers, 1): _hdr(ws1, 4, ci, h)
    ws1.row_dimensions[4].height = 30

    AGENT_KEYS = {"Agent","Due Date","Convert/Lost","Invalid Reason",
                  "Comments/Duplicate GRID","Feedback"}

    for ri, (_, row) in enumerate(df.iterrows(), DATA_S):
        lbl  = row["Label"]
        fill = _fill(lbl, ri)
        lc   = FONT_C.get(lbl,"000000")
        ws1.row_dimensions[ri].height = 20   # taller rows for Poppins 8pt

        for ci, key in enumerate(col_headers, 1):
            val = row.get(key,""); val = "" if pd.isna(val) else val

            # ── Agent workflow columns A–F: plain white, blank ────
            if key in AGENT_KEYS:
                c = ws1.cell(row=ri, column=ci, value="")
                c.border = _thin()
                c.font   = Font(name="Poppins", size=8)
                c.fill   = PatternFill("solid", start_color="FFFFFF")
                c.alignment = Alignment(vertical="center")
                if key == "Due Date":
                    c.number_format = "DD-MMM-YY"
                continue

            # ── Classifier output columns G onwards ───────────────
            c = ws1.cell(row=ri, column=ci, value=val)
            c.border = _thin()
            if key == "Label":
                c.font = Font(name="Poppins", size=8, bold=True, color=lc)
                c.fill = fill; c.alignment = Alignment(horizontal="center", vertical="center")
            elif key in ("Match Confidence","Match Score (CRM)"):
                try: c.value = float(val); c.number_format = "0.00"
                except Exception: pass
                c.font = Font(name="Poppins", size=8, bold=True)
                c.fill = _conf_fill(val)
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif key in ("Match Reason","Duplicate Match Method"):
                c.font = Font(name="Poppins", size=8, color="595959")
                c.fill = fill; c.alignment = Alignment(vertical="center", wrap_text=True)
            elif key == "Delivery Zone Status":
                zc = ("276221" if val == "Within Zone" else
                      "9C0006" if val == "Outside Zone" else "595959")
                zf = (PatternFill("solid", start_color="C6EFCE") if val == "Within Zone"
                      else PatternFill("solid", start_color="FFC7CE") if val == "Outside Zone"
                      else PatternFill("solid", start_color="EFEFEF"))
                c.font = Font(name="Poppins", size=8, bold=True, color=zc)
                c.fill = zf; c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.font = Font(name="Poppins", size=8)
                c.fill = fill; c.alignment = Alignment(vertical="center")

    for i, w in enumerate(col_w, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ── Dropdowns for agent columns ───────────────────────────────
    dv_convert = DataValidation(
        type="list", formula1='"Converted,Lost"', allow_blank=True)
    dv_reason  = DataValidation(
        type="list",
        formula1='"Duplicate,Invalid Data,Closed Down,Wrong Target Group,Other"',
        allow_blank=True)
    ws1.add_data_validation(dv_convert)
    ws1.add_data_validation(dv_reason)
    dv_convert.sqref = f"C{DATA_S}:C{DATA_E}"
    dv_reason.sqref  = f"D{DATA_S}:D{DATA_E}"

    ws1.freeze_panes     = "G5"
    ws1.auto_filter.ref  = f"A4:AI{DATA_E}"

    # ── Sheet 2: Summary ──────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Classification Summary"; ws2["A1"].font = Font(name="Poppins", bold=True, size=14, color="1F4E79")
    ws2["A2"] = "Formula-driven — auto-updates when Classified Leads is edited."
    ws2["A2"].font = Font(name="Poppins", italic=True, size=8, color="595959")
    ws2.merge_cells("A2:J2")

    r = 4
    _sec(ws2, r, 1, "CLASSIFICATION BREAKDOWN", 3); r += 1
    for ci, h in enumerate(["Label","Count","% of Total"], 1): _hdr(ws2, r, ci, h)
    r += 1
    s1 = r; total_r = s1 + len(labels_order)
    for i, lbl in enumerate(labels_order):
        ri = s1 + i
        _dc(ws2, ri, 1, lbl,  fill=FILLS.get(lbl), bold=True, align="left", color=FONT_C.get(lbl,"000000"))
        _dc(ws2, ri, 2, f'=COUNTIF({LBL_R},"{lbl}")', fill=FILLS.get(lbl))
        _dc(ws2, ri, 3, f'=IF(B{total_r}=0,0,B{ri}/B{total_r})', fill=FILLS.get(lbl), fmt="0.0%")
    r = total_r
    _dc(ws2, r, 1, "TOTAL", bold=True, align="left")
    _dc(ws2, r, 2, f'=SUM(B{s1}:B{total_r-1})', bold=True)
    _dc(ws2, r, 3, "100.0%", bold=True, fmt="0.0%")
    r += 2

    # Match method breakdown
    _sec(ws2, r, 1, "MATCH METHOD BREAKDOWN", 3); r += 1
    for ci, h in enumerate(["Method","Count","% of CRM Matches"], 1): _hdr(ws2, r, ci, h)
    r += 1
    crm_labels = ["P3 — Potential Match","P4 — Duplicate"]
    crm_total  = '+'.join([f'COUNTIF({LBL_R},"{l}")' for l in crm_labels])
    methods = [
        ("Postal+Unit+Name", f'=COUNTIFS({METH_R},"Postal+Unit*")', FILLS["P4 — Duplicate"]),
        ("Postal+Name",      f'=COUNTIFS({METH_R},"Postal+Name*")', ALT["P4 — Duplicate"]),
        ("New at known addr",f'=COUNTIFS({METH_R},"New business*")',FILLS["P3 — Potential Match"]),
    ]
    for ml, mf, mfill in methods:
        _dc(ws2, r, 1, ml, fill=mfill, align="left")
        _dc(ws2, r, 2, mf, fill=mfill)
        _dc(ws2, r, 3, f'=IF(({crm_total})=0,0,B{r}/({crm_total}))', fill=mfill, fmt="0.0%")
        r += 1
    r += 1

    # Top GM categories
    _sec(ws2, r, 1, "TOP GM CATEGORIES (Qualified Leads)", 3); r += 1
    for ci, h in enumerate(["Category","Count","% of Matched"], 1): _hdr(ws2, r, ci, h)
    r += 1
    top_cats = (df[df["GM Category"].notna() & (df["GM Category"] != "")]
                ["GM Category"].value_counts().head(20).index.tolist())
    matched_f = f'=COUNTIF({LBL_R},"P1 — New")'
    hid_r = r + len(top_cats) + 1
    ws2.cell(row=hid_r, column=2).value = matched_f
    ws2.cell(row=hid_r, column=2).font  = Font(color="FFFFFF", size=1)
    for i, cat in enumerate(top_cats):
        fl = (PatternFill("solid", start_color="EBF3FB")
              if i % 2 == 0 else PatternFill("solid", start_color="FFFFFF"))
        _dc(ws2, r, 1, cat, fill=fl, align="left")
        _dc(ws2, r, 2, f'=COUNTIF({CAT_R},"{cat}")', fill=fl)
        _dc(ws2, r, 3, f'=IF(B{hid_r}=0,0,B{r}/B{hid_r})', fill=fl, fmt="0.0%")
        r += 1
    r = hid_r + 2

    # Zone section
    if has_z:
        _sec(ws2, r, 1, "DELIVERY ZONE BREAKDOWN", 3); r += 1
        for ci, h in enumerate(["Zone Status","Count","% of Total"], 1): _hdr(ws2, r, ci, h)
        r += 1
        for zs, zcol, bg in [("Within Zone","276221","C6EFCE"),
                               ("Outside Zone","9C0006","FFC7CE"),
                               ("Geocoding Failed","595959","D9D9D9")]:
            fl = PatternFill("solid", start_color=bg)
            _dc(ws2, r, 1, zs, fill=fl, bold=True, align="left", color=zcol)
            _dc(ws2, r, 2, f'=COUNTIF({ZONE_R},"{zs}")', fill=fl)
            _dc(ws2, r, 3, f'=IF(B{total_r}=0,0,B{r}/B{total_r})', fill=fl, fmt="0.0%")
            r += 1

    for col, w in zip("ABCDE",[32,12,14,4,26]):
        ws2.column_dimensions[col].width = w

    # ── Sheet 3: P1 New ───────────────────────────────────────────
    def _mini_sheet(wb, title_text, title_color, filter_fn, heads, widths, sheet_name):
        ws = wb.create_sheet(sheet_name)
        sub = df[filter_fn(df)].reset_index(drop=True)
        ws["A1"] = f"{title_text}  ({len(sub):,} leads)"
        ws["A1"].font = Font(name="Poppins", bold=True, size=12, color=title_color)
        ws.merge_cells(f"A1:{get_column_letter(len(heads))}1")
        for ci, h in enumerate(heads, 1): _hdr(ws, 3, ci, h)
        for ri, (_, row) in enumerate(sub.iterrows(), 4):
            lbl  = row["Label"]
            fill = _fill(lbl, ri)
            lc   = FONT_C.get(lbl,"000000")
            ws.row_dimensions[ri].height = 20
            for ci, key in enumerate(heads, 1):
                val = row.get(key,""); val = "" if pd.isna(val) else val
                c   = ws.cell(row=ri, column=ci, value=val)
                c.border = _thin()
                if key == "Label":
                    c.font = Font(name="Poppins",size=8,bold=True,color=lc)
                    c.fill = fill; c.alignment = Alignment(horizontal="center",vertical="center")
                elif key in ("Match Confidence","Match Score (CRM)"):
                    try: c.value = float(val); c.number_format = "0.00"
                    except Exception: pass
                    c.font = Font(name="Poppins",size=8,bold=True)
                    c.fill = _conf_fill(val)
                    c.alignment = Alignment(horizontal="center",vertical="center")
                elif key in ("Match Reason","Duplicate Match Method"):
                    c.font = Font(name="Poppins",size=8,color="595959")
                    c.fill = fill; c.alignment = Alignment(vertical="center",wrap_text=True)
                else:
                    c.font = Font(name="Poppins",size=8)
                    c.fill = fill; c.alignment = Alignment(vertical="center")
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A4"
        if len(sub) > 0:
            ws.auto_filter.ref = f"A3:{get_column_letter(len(heads))}{3+len(sub)}"
        return ws

    new_h = ["GRID","Lead ID","Company / Account","City","Street","Phone",
             "GM Title","GM Category","GM Business Status",
             "GM Phone","GM Website","GM URL","Match Confidence","Match Reason",
             "Previous Occupant GRID","Previous Occupant Name",
             "Delivery Zone Status","Zone Name"]
    _mini_sheet(wb, "✅ P1 — New Leads", "276221",
                lambda d: d["Label"] == "P1 — New",
                new_h, [10,18,32,12,36,16,28,24,18,16,28,48,14,38,14,30,18,18],
                "✅ P1 — New")

    dup_h = ["GRID","Lead ID","Company / Account","City","Street","Phone",
             "Label","Duplicate GRID","Duplicate CRM Name",
             "Duplicate CRM Address","Duplicate CRM Postal",
             "CRM Account Status","CRM Status Reason",
             "Duplicate Match Method","Match Score (CRM)"]
    _mini_sheet(wb, "🔴 P4 — Duplicates", "9C0006",
                lambda d: d["Label"] == "P4 — Duplicate",
                dup_h, [10,18,32,12,36,16,20,14,30,36,14,16,18,40,14],
                "🔴 P4 — Duplicate")

    pot_h = ["GRID","Lead ID","Company / Account","City","Street","Phone",
             "Label","Duplicate GRID","Duplicate CRM Name",
             "Duplicate CRM Address","Duplicate CRM Postal",
             "CRM Account Status","Duplicate Match Method","Match Score (CRM)",
             "GM Title","GM Category","Match Confidence","Match Reason"]
    _mini_sheet(wb, "🟡 P3 — Potential Match", "7D5A00",
                lambda d: d["Label"] == "P3 — Potential Match",
                pot_h, [10,18,32,12,36,16,22,14,30,36,14,16,40,14,28,24,14,38],
                "🟡 P3 — Potential")

    chk_h = ["GRID","Lead ID","Company / Account","City","Street",
             "GM Title","GM Category","GM Business Status",
             "Match Confidence","Match Reason","GM URL"]
    _mini_sheet(wb, "⚪ P2 — Please Check", "595959",
                lambda d: d["Label"] == "P2 — Please Check",
                chk_h, [10,18,32,12,36,28,24,18,14,38,48],
                "⚪ P2 — Please Check")

    other_h = ["GRID","Lead ID","Company / Account","City","Phone","Label",
               "GM Title","GM Category","GM Business Status",
               "Match Confidence","Match Reason","GM URL"]
    _mini_sheet(wb, "⚠️ Closed / Wrong Target Group", "7D4E00",
                lambda d: d["Label"].isin(["Business Closed","Wrong Target Group"]),
                other_h, [10,18,32,12,16,22,28,24,18,14,38,48],
                "⚠️ Closed + Wrong TG")

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf


# ═════════════════════════════════════════════════════════════════
# STREAMLIT MAIN
# ═════════════════════════════════════════════════════════════════

def _col_map_ui(df, auto_map: dict, fields: list, key_prefix: str) -> dict:
    """
    Render a column mapping expander.
    fields: list of (internal_key, display_label, required: bool)
    Returns updated mapping dict {key: column_name_or_None}.
    Auto-expands if any required field is unmapped.
    """
    all_cols  = ["(not mapped)"] + list(df.columns)
    missing   = [label for key, label, req in fields
                 if req and auto_map.get(key) is None]
    expanded  = bool(missing)
    icon      = "⚠️" if missing else "✅"
    label_txt = (f"Column mapping {icon} — missing: {', '.join(missing)}"
                 if missing else f"Column mapping {icon} — all detected")

    result = dict(auto_map)
    with st.expander(label_txt, expanded=expanded):
        cols_ui = st.columns(2)
        for idx, (key, label, _) in enumerate(fields):
            current = auto_map.get(key)
            default = all_cols.index(current) if current and current in all_cols else 0
            sel = cols_ui[idx % 2].selectbox(
                label, all_cols, index=default,
                key=f"{key_prefix}_{key}")
            result[key] = sel if sel != "(not mapped)" else None
    return result


# ── Google Maps → Salesforce cuisine picklist mapping ─────────────
APIFY_CUISINE_MAP = {
    "thai restaurant":"Thai","thai food":"Thai",
    "japanese restaurant":"Japanese","ramen restaurant":"Japanese","izakaya":"Japanese",
    "sushi restaurant":"Sushi",
    "korean restaurant":"Korean",
    "chinese restaurant":"Chinese",
    "indian restaurant":"Indian",
    "italian restaurant":"Italian",
    "vietnamese restaurant":"Vietnamese",
    "pizza restaurant":"Pizza","pizza delivery":"Pizza",
    "malay restaurant":"Malay","nasi padang restaurant":"Malay",
    "indonesian restaurant":"Indonesian",
    "western restaurant":"Western",
    "fast food restaurant":"Fast Food",
    "seafood restaurant":"Seafood",
    "vegetarian restaurant":"Vegetarian","vegan restaurant":"Vegetarian",
    "american restaurant":"American",
    "french restaurant":"French","patisserie":"French",
    "mexican restaurant":"Mexican",
    "mediterranean restaurant":"Mediterranean",
    "middle eastern restaurant":"Middle Eastern",
    "filipino restaurant":"Filipino",
    "sandwich shop":"Sandwiches","sandwich restaurant":"Sandwiches",
    "german restaurant":"German",
    "greek restaurant":"Greek",
    "spanish restaurant":"Spanish",
    "turkish restaurant":"Turkish",
    "lebanese restaurant":"Lebanese",
    "malaysian restaurant":"Malaysian",
    "bakery":"Cakes & Bakery","cake shop":"Cakes & Bakery","confectionery":"Cakes & Bakery",
    "dessert shop":"Desserts","dessert restaurant":"Desserts","ice cream shop":"Desserts",
    "bubble tea shop":"Non-alcoholic Drinks","juice bar":"Non-alcoholic Drinks",
    "tea house":"Non-alcoholic Drinks",
    "chicken restaurant":"Chicken","fried chicken restaurant":"Chicken",
    "burger restaurant":"Burgers","hamburger restaurant":"Burgers",
    "steakhouse":"Meat","barbecue restaurant":"Meat",
    "halal restaurant":"Halal",
    "health food restaurant":"Healthy Food","salad shop":"Healthy Food",
    "singaporean restaurant":"Singaporean","hawker stall":"Singaporean",
    "asian restaurant":"Asian","pan asian restaurant":"Asian",
    "southeast asian restaurant":"South East Asian",
    "international restaurant":"International",
}

_NAMING_RE = re.compile(r'^(.+?)\s*\(([^)]+)\)\s*$')


def _map_apify_cuisine(category: str) -> str:
    """Map Google Maps category to Salesforce cuisine picklist. Returns '' if unmappable."""
    if not category:
        return ""
    cat = str(category).strip().lower()
    if cat in APIFY_CUISINE_MAP:
        return APIFY_CUISINE_MAP[cat]
    for key, val in APIFY_CUISINE_MAP.items():
        if key in cat:
            return val
    return ""


def _check_naming_convention(name: str):
    """Check RESTAURANT NAME (LOCATION) structure only.
    Valid: any text followed by (any text in brackets).
    No case enforcement — Title Case and ALL CAPS are both acceptable.
    Returns (ok: bool, detail: str)."""
    if not name or (isinstance(name, float) and pd.isna(name)):
        return False, "Account name is blank"
    name = str(name).strip()
    m = _NAMING_RE.match(name)
    if not m:
        return False, "Missing (Location) part — expected: NAME (LOCATION)"
    if not m.group(1).strip():
        return False, "Restaurant name part is empty"
    if not m.group(2).strip():
        return False, "Location part inside brackets is empty"
    return True, "Pass"


def clean_lead_report(df: pd.DataFrame) -> pd.DataFrame:
    """Strip Salesforce footer rows and rows with invalid GRID."""
    FOOTER = ["confidential", "copyright", "salesforce.com"]
    mask = df["Lead Source"].astype(str).str.lower().apply(
        lambda v: not any(f in v for f in FOOTER))
    df = df[mask & df["GRID"].astype(str).str.strip().ne("nan")].copy()
    return df.reset_index(drop=True)


def sample_leads(df: pd.DataFrame, seed: int = 42) -> tuple:
    """
    Stratified 10% sample by Edited By × Lead Source × Lead Status.
    Returns (sampled_df, summary_df).
    """
    import math
    strata = ["Edited By", "Lead Source", "Lead Status"]
    sampled, summary = [], []
    for keys, grp in df.groupby(strata, dropna=False):
        n = len(grp)
        if n == 0:
            continue
        k = min(max(1, math.ceil(n * 0.10)), n)
        sampled.append(grp.sample(n=k, random_state=seed))
        summary.append({
            "Edited By": keys[0], "Lead Source": keys[1],
            "Lead Status": keys[2], "Total Leads": n, "Sampled": k,
        })
    sampled_df  = pd.concat(sampled,  ignore_index=True) if sampled  else df.iloc[:0]
    summary_df  = pd.DataFrame(summary)
    return sampled_df, summary_df


def run_kpi_checks(sampled_df, apify_df, crm_df, col_map_crm,
                   account_df, account_cols,
                   zones, char_map,
                   p2_threshold=0.50, p3_threshold=0.75):
    """
    Run all KPI checks on sampled leads.
    Returns (results_df, agent_summary_df).
    """

    def _sv(row, col, default=""):
        """Safe scalar from a pandas Series row.
        Handles duplicate column names, pd.NA, NaT, and None."""
        if not col or row is None:
            return default
        try:
            if col not in row.index:
                return default
            val = row[col]
            if isinstance(val, pd.Series):   # duplicate column → take first
                val = val.iloc[0] if len(val) > 0 else default
            if pd.isna(val):
                return default
            return val
        except Exception:
            return default

    # ── Build CRM indexes ─────────────────────────────────────────
    crm_postal_all_dict  = {}
    crm_postal_unit_dict = {}
    all_crm_items        = []
    postal_c = col_map_crm.get("postal") if col_map_crm else None
    addr_c   = col_map_crm.get("street") if col_map_crm else None
    name_c   = col_map_crm.get("name")   if col_map_crm else None
    grid_c   = col_map_crm.get("grid")   if col_map_crm else None

    if crm_df is not None:
        for _, r in crm_df.iterrows():
            raw_n      = str(_sv(r, name_c)   or "")
            lat_n, pin_n = norm_name_sg(strip_venue_generic(raw_n), char_map)
            item       = (raw_n, lat_n, pin_n, r)
            all_crm_items.append(item)
            postal_raw = _norm_postal_input(_sv(r, postal_c))
            addr_raw   = str(_sv(r, addr_c)   or "")
            crm_p  = extract_sg_postal(postal_raw) or extract_sg_postal(addr_raw)
            crm_u  = extract_sg_unit(addr_raw)
            if crm_p:
                crm_postal_all_dict.setdefault(crm_p, []).append(item)
                if crm_u:
                    crm_postal_unit_dict.setdefault(f"{crm_p}|{crm_u}", []).append(item)

    # ── Build Apify dict (GRID → row) ─────────────────────────────
    apify_dict = {}
    if apify_df is not None:
        gc = detect_column(apify_df, ["GRID","grid","Grid"])
        if gc:
            for _, r in apify_df.iterrows():
                g = str(_sv(r, gc) or "").strip()
                if g and g not in apify_dict:
                    apify_dict[g] = r
        _apy_title = detect_column(apify_df, ["title","name"])        or "title"
        _apy_cat   = detect_column(apify_df, ["categoryName","category"]) or "categoryName"
        _apy_perm  = detect_column(apify_df, ["permanentlyClosed"])   or "permanentlyClosed"
        _apy_temp  = detect_column(apify_df, ["temporarilyClosed"])   or "temporarilyClosed"
        _apy_phone = detect_column(apify_df, ["phone","phoneNumber"]) or "phone"
        _apy_web   = detect_column(apify_df, ["website"])             or "website"
    else:
        _apy_title = _apy_cat = _apy_perm = _apy_temp = _apy_phone = _apy_web = ""

    # ── Build account details dict (GRID → row) ───────────────────
    account_dict = {}
    if account_df is not None:
        ag = account_cols.get("grid")
        if ag:
            for _, r in account_df.iterrows():
                g = str(_sv(r, ag) or "").strip()
                if g and g not in account_dict:
                    account_dict[g] = r

    # ── Scoring function ──────────────────────────────────────────
    def _sc(lead_lat, lead_pin, item):
        raw_n, lat_n, pin_n, r = item
        ns = 0.0
        if lead_lat and lat_n:
            ns = max(fuzz.token_sort_ratio(lead_lat, lat_n),
                     fuzz.token_set_ratio(lead_lat, lat_n)) / 100.0
        if lead_pin or pin_n:
            lp, cp = lead_pin or "", pin_n or ""
            if lp and cp:
                ps = max(fuzz.token_sort_ratio(lp, cp),
                         fuzz.token_set_ratio(lp, cp)) / 100.0
                return round(max(ns, ps), 3), raw_n, r
        return round(ns, 3), raw_n, r

    # ── Dedup helper (excludes own GRID) ─────────────────────────
    NO_UNIT_MIN = max(p2_threshold, 0.90)

    def _find_dup(lead_grid, lead_lat, lead_pin, lead_postal, lead_unit):
        def _excl(pool):
            return [i for i in pool
                    if str(_sv(i[3], grid_c) or "") != lead_grid]
        best_sc, best_row, best_raw = 0.0, None, ""
        if lead_postal:
            if lead_unit:
                um = crm_postal_unit_dict.get(f"{lead_postal}|{lead_unit}", [])
                pool   = _excl(um) if um else _excl(crm_postal_all_dict.get(lead_postal,[]))
                min_sc = p2_threshold if um else NO_UNIT_MIN
            else:
                pool   = _excl(crm_postal_all_dict.get(lead_postal, []))
                min_sc = NO_UNIT_MIN
            for item in pool:
                sc, raw_n, cand = _sc(lead_lat, lead_pin, item)
                if sc > best_sc:
                    best_sc, best_row, best_raw = sc, cand, raw_n
            if best_row is not None and best_sc >= min_sc:
                return best_row, best_sc, f"Postal+Name [{lead_postal}] score={best_sc:.2f}"
        else:
            for item in _excl(all_crm_items):
                sc, raw_n, cand = _sc(lead_lat, lead_pin, item)
                if sc >= 0.90 and sc > best_sc:
                    best_sc, best_row, best_raw = sc, cand, raw_n
            if best_row is not None:
                return best_row, best_sc, f"Name-only (zero postal) score={best_sc:.2f}"
        return None, 0.0, ""

    results = []

    for _, row in sampled_df.iterrows():
        lead_grid   = str(row.get("GRID","") or "").strip()
        lead_src    = str(row.get("Lead Source","") or "")
        lead_stat   = str(row.get("Lead Status","") or "")
        lost_rsn    = str(row.get("Lost Reason","") or "") \
                      if pd.notna(row.get("Lost Reason")) else ""
        edited_by   = str(row.get("Edited By","") or "")
        edit_date   = row.get("Edit Date")
        street      = str(row.get("Street","") or "")
        postal_raw  = str(row.get("Zip/Postal Code","") or "")
        dup_id_agent = str(row.get("Duplicate Id","") or "") \
                       if pd.notna(row.get("Duplicate Id")) else ""

        lead_postal = extract_sg_postal(_norm_postal_input(postal_raw)) \
                      or extract_sg_postal(street)
        lead_unit   = extract_sg_unit(street)

        try:
            edit_dt = pd.to_datetime(str(edit_date), dayfirst=True, errors="coerce")
        except Exception:
            edit_dt = None

        # ── Safe scalar extractor (alias for _sv defined above) ───
        _gs = _sv

        # ── Apify row ──────────────────────────────────────────────
        apy = apify_dict.get(lead_grid)
        gm_title = gm_cat = gm_biz_status = gm_phone = gm_web = ""
        perm = temp = False
        if apy is not None:
            gm_title = str(_gs(apy, _apy_title, ""))
            gm_cat   = str(_gs(apy, _apy_cat,   ""))
            gm_phone = str(_gs(apy, _apy_phone,  ""))
            gm_web   = str(_gs(apy, _apy_web,    ""))
            perm_v   = _gs(apy, _apy_perm, False)
            temp_v   = _gs(apy, _apy_temp, False)
            perm     = str(perm_v).lower() in ("true","1","yes")
            temp     = str(temp_v).lower() in ("true","1","yes")
            gm_biz_status = ("Permanently Closed" if perm else
                             "Temporarily Closed"  if temp else "Open")

        # ── Account details (converted only) ──────────────────────
        acc = account_dict.get(lead_grid) if lead_stat == "Converted" else None

        # Lead name: from account (converted) or Apify title
        an_col    = account_cols.get("name") if account_cols else None
        lead_name = (str(_gs(acc, an_col) or "") if acc is not None and an_col else "") or gm_title
        lead_lat_n, lead_pin_n = norm_name_sg(strip_venue_generic(lead_name), char_map)

        # Restaurant name for output:
        # Converted → Account Name from Account Details report
        # Lost      → Company column from lead report
        if lead_stat == "Converted":
            restaurant_name = str(_gs(acc, an_col) or "") if acc is not None and an_col else ""
        else:
            # Try both column names (XLS vs CSV export formats)
            restaurant_name = (str(row.get("Company","") or "")
                               or str(row.get("Company / Account","") or ""))

        # ── Run checks ────────────────────────────────────────────
        C = {}

        # C01 — Valid restaurant
        if apy is None:
            C["C01 Valid Restaurant"] = "⚠️ No Apify result — unverifiable"
        elif perm or temp:
            C["C01 Valid Restaurant"] = \
                f"❌ {'Permanently' if perm else 'Temporarily'} closed on Google Maps"
        elif not gm_cat:
            C["C01 Valid Restaurant"] = "⚠️ No category in Apify — unverifiable"
        elif not is_food_delivery_eligible(gm_cat, _DEFAULT_EXCLUSION_KW):
            C["C01 Valid Restaurant"] = f"❌ Non-food category: {gm_cat}"
        else:
            C["C01 Valid Restaurant"] = f"✅ Pass — {gm_cat}"

        # C03 — Dedup (needed for C02 and C05)
        dup_row, dup_sc, dup_meth = _find_dup(
            lead_grid, lead_lat_n, lead_pin_n, lead_postal, lead_unit)
        dup_name_found  = str(_gs(dup_row, name_c) or "") if dup_row is not None and name_c else ""
        dup_grid_found  = str(_gs(dup_row, grid_c) or "") if dup_row is not None and grid_c else ""
        dup_created_str = ""
        if dup_row is not None:
            for dc in ["Created Date","created_date","CreatedDate"]:
                v = _gs(dup_row, dc)
                if v and str(v) not in ("","nan"):
                    dup_created_str = str(v); break

        if not lead_name:
            C["C03 Duplicate"] = "⚠️ No lead name — dedup unverifiable"
        elif dup_row is not None:
            C["C03 Duplicate"] = \
                f"⚠️ Potential duplicate: {dup_name_found} ({dup_grid_found}) score={dup_sc:.2f}"
        else:
            C["C03 Duplicate"] = "✅ No duplicate found"

        # C02 — Lost reason (lost only)
        if lead_stat != "Lost":
            C["C02 Lost Reason"] = "N/A — Converted"
        elif lost_rsn == "Other":
            C["C02 Lost Reason"] = "⚠️ Unverifiable — likely test account"
        elif lost_rsn == "Closed Down":
            if apy is None:
                C["C02 Lost Reason"] = "⚠️ No Apify result — unverifiable"
            elif perm or temp:
                C["C02 Lost Reason"] = "✅ Pass — Google confirms closed"
            else:
                C["C02 Lost Reason"] = f"❌ Restaurant appears open on Google ({gm_biz_status})"
        elif lost_rsn == "Wrong Target Group":
            if apy is None:
                C["C02 Lost Reason"] = "⚠️ No Apify result — unverifiable"
            elif not gm_cat:
                C["C02 Lost Reason"] = "⚠️ No Apify category — unverifiable"
            elif not is_food_delivery_eligible(gm_cat, _DEFAULT_EXCLUSION_KW):
                C["C02 Lost Reason"] = f"✅ Pass — confirmed non-food: {gm_cat}"
            else:
                C["C02 Lost Reason"] = f"❌ Category appears food-eligible: {gm_cat}"
        elif lost_rsn == "Invalid Data":
            if apy is None:
                C["C02 Lost Reason"] = "✅ Pass — no Google listing found (flag for manual check)"
            else:
                C["C02 Lost Reason"] = \
                    f"⚠️ Has Google listing ({gm_title}) — verify if truly no online presence"
        elif lost_rsn == "No Delivery Service":
            if not lead_postal:
                C["C02 Lost Reason"] = "⚠️ No postal — zone check unverifiable"
            elif not zones:
                C["C02 Lost Reason"] = f"⚠️ No zone data — verify postal {lead_postal} manually"
            else:
                # Try to check via geocoding if available
                try:
                    token = get_onemap_token()
                    coords = geocode_postal_sg(lead_postal, token) if token else None
                    if coords:
                        lat, lng = coords
                        z_stat, z_name, z_city, _ = classify_zone(lat, lng, zones)
                        if z_stat == "Outside Zone":
                            C["C02 Lost Reason"] = f"✅ Pass — outside delivery zones"
                        else:
                            C["C02 Lost Reason"] = \
                                f"❌ Postal {lead_postal} is within zone '{z_name}'"
                    else:
                        C["C02 Lost Reason"] = \
                            f"⚠️ Could not geocode postal {lead_postal} — verify manually"
                except Exception:
                    C["C02 Lost Reason"] = \
                        f"⚠️ Zone check failed — verify postal {lead_postal} manually"
        elif lost_rsn == "Duplicate":
            if dup_row is not None:
                C["C02 Lost Reason"] = \
                    f"✅ Pass — duplicate confirmed: {dup_name_found} ({dup_grid_found})"
                # Update C03
                C["C03 Duplicate"] = \
                    f"✅ Confirmed duplicate: {dup_name_found} ({dup_grid_found}) score={dup_sc:.2f}"
            else:
                C["C02 Lost Reason"] = "❌ Lost as Duplicate but no duplicate found by tool"
                C["C03 Duplicate"]   = "❌ No duplicate found — lost reason may be incorrect"
        else:
            C["C02 Lost Reason"] = f"⚠️ Unknown lost reason: {lost_rsn}"

        # C04 — Agent's Duplicate ID accuracy
        if lead_stat != "Lost" or lost_rsn != "Duplicate":
            C["C04 Duplicate ID Accuracy"] = "N/A"
        elif not dup_id_agent:
            C["C04 Duplicate ID Accuracy"] = "❌ Agent did not enter Duplicate ID"
        elif dup_grid_found and dup_id_agent.strip() == dup_grid_found.strip():
            C["C04 Duplicate ID Accuracy"] = f"✅ Match — agent: {dup_id_agent}"
        elif dup_grid_found:
            C["C04 Duplicate ID Accuracy"] = \
                f"⚠️ Mismatch — agent: {dup_id_agent} | tool: {dup_grid_found} ({dup_name_found})"
        else:
            C["C04 Duplicate ID Accuracy"] = \
                f"⚠️ No tool match to compare — agent entered: {dup_id_agent}"

        # C05 — Wrongful conversion
        if lead_stat != "Converted":
            C["C05 Wrongful Conversion"] = "N/A — Lost lead"
        elif dup_row is None:
            C["C05 Wrongful Conversion"] = "✅ No duplicate found"
        else:
            try:
                dup_created_dt = pd.to_datetime(dup_created_str, dayfirst=True, errors="coerce")
            except Exception:
                dup_created_dt = None
            if dup_created_dt is None or edit_dt is None:
                C["C05 Wrongful Conversion"] = \
                    f"⚠️ Duplicate found ({dup_name_found}) — cannot compare dates, manual review"
            elif dup_created_dt.date() <= edit_dt.date():
                C["C05 Wrongful Conversion"] = \
                    f"❌ Duplicate ({dup_name_found}) existed before conversion " \
                    f"[dup created: {dup_created_dt.date()}, lead converted: {edit_dt.date()}]"
            else:
                C["C05 Wrongful Conversion"] = \
                    f"✅ Duplicate created after conversion — not wrongful " \
                    f"[dup: {dup_created_dt.date()}, converted: {edit_dt.date()}]"

        # C06 — Naming convention (converted only)
        if lead_stat != "Converted":
            C["C06 Naming Convention"] = "N/A"
        elif not lead_name:
            C["C06 Naming Convention"] = "⚠️ No account name available"
        else:
            ok, detail = _check_naming_convention(lead_name)
            C["C06 Naming Convention"] = "✅ Pass" if ok else f"❌ {detail}"

        # Helpers for account detail checks
        def _acc_check(key, label):
            """Returns check string for a simple populated/blank field."""
            if lead_stat != "Converted": return "N/A"
            if acc is None: return "⚠️ No account details uploaded"
            col = account_cols.get(key)
            val = str(_gs(acc, col) or "") if col else ""
            if not val or val.lower() == "nan":
                return f"❌ {label} not populated"
            return f"✅ Populated: {val}"

        # C07 — Phone
        if lead_stat != "Converted":
            C["C07 Phone"] = "N/A"
        elif acc is None:
            C["C07 Phone"] = "⚠️ No account details uploaded"
        else:
            ph_col  = account_cols.get("phone")
            sf_ph   = str(_gs(acc, ph_col) or "") if ph_col else ""
            if not sf_ph or sf_ph.lower() == "nan":
                C["C07 Phone"] = "❌ Phone not populated in Salesforce"
            elif gm_phone:
                sf_n = re.sub(r'\D','', sf_ph)
                gm_n = re.sub(r'\D','', gm_phone)
                if sf_n == gm_n:
                    C["C07 Phone"] = f"✅ Match — {sf_ph}"
                else:
                    C["C07 Phone"] = \
                        f"⚠️ Mismatch — SF: {sf_ph} | Google: {gm_phone}"
            else:
                C["C07 Phone"] = f"✅ Populated: {sf_ph} (no Apify phone to compare)"

        C["C08 Email"]          = _acc_check("email",          "Email")
        C["C10 Social Media"]   = _acc_check("social_media",   "Social Media URL")
        C["C11 Parent Account"] = _acc_check("parent_account", "Parent Account")
        C["C12 Business Office"]= _acc_check("business_office","Business Office")
        C["C14 Target Partner"] = _acc_check("target_partner", "Target Partner")

        # C09 — Website
        if lead_stat != "Converted":
            C["C09 Website"] = "N/A"
        elif acc is None:
            C["C09 Website"] = "⚠️ No account details uploaded"
        else:
            ws_col = account_cols.get("website")
            sf_ws  = str(_gs(acc, ws_col) or "") if ws_col else ""
            if not sf_ws or sf_ws.lower() == "nan":
                C["C09 Website"] = "❌ Website not populated"
            elif gm_web:
                def _nu(u):
                    return re.sub(r'^https?://(www\.)?','',
                                  str(u).lower().strip()).rstrip('/')
                if _nu(sf_ws) == _nu(gm_web):
                    C["C09 Website"] = f"✅ Match — {sf_ws}"
                else:
                    C["C09 Website"] = \
                        f"⚠️ Mismatch — SF: {sf_ws} | Google: {gm_web}"
            else:
                C["C09 Website"] = f"✅ Populated: {sf_ws}"

        # C13 — Delivery service
        if lead_stat != "Converted":
            C["C13 Delivery Service"] = "N/A"
        elif acc is None:
            C["C13 Delivery Service"] = "⚠️ No account details uploaded"
        else:
            ds_col = account_cols.get("delivery_service")
            sf_ds  = str(_gs(acc, ds_col) or "") if ds_col else ""
            if not sf_ds or sf_ds.lower() == "nan":
                C["C13 Delivery Service"] = "❌ Delivery Service not populated"
            else:
                dsl  = sf_ds.lower()
                h_dh = "dh delivery" in dsl
                h_ta = "take away" in dsl or "takeaway" in dsl
                if h_dh and h_ta:
                    C["C13 Delivery Service"] = f"✅ Pass — {sf_ds}"
                else:
                    missing = []
                    if not h_dh: missing.append("DH Delivery")
                    if not h_ta: missing.append("Take Away")
                    C["C13 Delivery Service"] = \
                        f"❌ Missing: {', '.join(missing)} — current: {sf_ds}"

        # C15 — Restaurant category
        if lead_stat != "Converted":
            C["C15 Restaurant Category"] = "N/A"
        elif acc is None:
            C["C15 Restaurant Category"] = "⚠️ No account details uploaded"
        else:
            cat_col = account_cols.get("category")
            sf_cat  = str(_gs(acc, cat_col) or "") if cat_col else ""
            if not sf_cat or sf_cat.lower() == "nan":
                C["C15 Restaurant Category"] = "❌ Category not populated"
            elif not gm_cat:
                C["C15 Restaurant Category"] = \
                    f"⚠️ Cannot verify — no Apify category. SF: {sf_cat}"
            else:
                mapped = _map_apify_cuisine(gm_cat)
                if not mapped:
                    C["C15 Restaurant Category"] = \
                        f"⚠️ Apify category '{gm_cat}' unmappable — SF: {sf_cat}"
                elif mapped.lower() == sf_cat.lower():
                    C["C15 Restaurant Category"] = f"✅ Match — {sf_cat}"
                else:
                    C["C15 Restaurant Category"] = \
                        f"⚠️ Possible mismatch — SF: {sf_cat} | Apify suggests: {mapped}"

        auto_errs = sum(1 for v in C.values() if str(v).startswith("❌"))

        results.append({
            "GRID":              lead_grid,
            "Restaurant Name":   restaurant_name,
            "Lead Source":       lead_src,
            "Lead Status":     lead_stat,
            "Lost Reason":     lost_rsn,
            "Edited By":       edited_by,
            "Edit Date":       edit_date,
            "Street":          street,
            "Zip/Postal Code": _norm_postal_input(postal_raw) or str(postal_raw).replace(".0",""),
            "GM Title":        gm_title,
            "GM Category":     gm_cat,
            "GM Status":       gm_biz_status,
            **C,
            "Auto Error Count":      auto_errs,
            "Ops Feedback":          "",
            "Updated Error Count":   "",
            "Agent Acknowledgement": "",
            "Agent Feedback":        "",
        })

    results_df = pd.DataFrame(results)

    # Agent summary
    agg = results_df.groupby("Edited By", dropna=False).agg(
        Leads_Sampled=("GRID","count"),
        Auto_Errors=("Auto Error Count","sum")
    ).reset_index()
    agg.columns = ["Agent","Leads Sampled","Auto Error Count"]
    agg["Updated Error Count"] = ""

    return results_df, agg


def build_kpi_excel(results_df: pd.DataFrame,
                    agent_df: pd.DataFrame,
                    sampling_df: pd.DataFrame) -> bytes:
    """Build KPI scorecard Excel with 3 sheets."""
    import io as _io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    FNT = "Poppins"

    CHECK_COLS = [c for c in results_df.columns
                  if c.startswith("C") and c[1:3].isdigit()]
    AGENT_COLS_OUT = [
        "Auto Error Count","Ops Feedback","Updated Error Count",
        "Agent Acknowledgement","Agent Feedback",
    ]
    INFO_COLS = ["GRID","Restaurant Name","Lead Source","Lead Status","Lost Reason",
                 "Edited By","Edit Date","Street","Zip/Postal Code",
                 "GM Title","GM Category","GM Status"]
    ALL_COLS  = INFO_COLS + CHECK_COLS + AGENT_COLS_OUT

    # Pre-compute check column letter range for COUNTIF formula
    _chk_s = get_column_letter(len(INFO_COLS) + 1)
    _chk_e = get_column_letter(len(INFO_COLS) + len(CHECK_COLS))

    def _thin():
        s = Side(style="thin", color="D0D0D0")
        return Border(left=s, right=s, top=s, bottom=s)

    HDR = PatternFill("solid", start_color="1A1A2E")
    AGT = PatternFill("solid", start_color="EEF2FF")

    FILLS = {
        "✅": PatternFill("solid", start_color="C6EFCE"),
        "❌": PatternFill("solid", start_color="FFC7CE"),
        "⚠️": PatternFill("solid", start_color="FFF2CC"),
        "N/A": PatternFill("solid", start_color="F0F0F0"),
    }
    FONTS = {
        "✅":"276221","❌":"9C0006","⚠️":"7D5A00","N/A":"595959",
    }

    def _cell_fill(val):
        v = str(val)
        for k in ["✅","❌","⚠️","N/A"]:
            if v.startswith(k):
                return FILLS[k], FONTS[k]
        return PatternFill(), "000000"

    # ── Sheet 1: Sampled Leads ─────────────────────────────────────
    ws1 = wb.active; ws1.title = "Sampled Leads"
    ws1["A1"] = f"KPI Sample Checker — {pd.Timestamp.now().strftime('%B %Y')}"
    ws1["A1"].font = Font(name=FNT, bold=True, size=14, color="1F4E79")
    ws1.merge_cells(f"A1:{get_column_letter(len(ALL_COLS))}1")
    ws1.row_dimensions[1].height = 24

    for ci, h in enumerate(ALL_COLS, 1):
        c = ws1.cell(row=2, column=ci, value=h)
        c.font      = Font(name=FNT, bold=True, color="FFFFFF", size=9)
        c.fill      = HDR
        c.border    = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[2].height = 30

    for ri, (_, row) in enumerate(results_df[ALL_COLS].iterrows(), 3):
        ws1.row_dimensions[ri].height = 18
        for ci, col in enumerate(ALL_COLS, 1):
            val = row[col]; val = "" if pd.isna(val) else val
            c   = ws1.cell(row=ri, column=ci, value=val)
            c.border = _thin()
            c.font   = Font(name=FNT, size=8)

            if col in CHECK_COLS:
                fill, fc = _cell_fill(val)
                c.fill      = fill
                c.font      = Font(name=FNT, size=8, color=fc)
                c.alignment = Alignment(vertical="center", wrap_text=True)
            elif col in AGENT_COLS_OUT:
                c.fill      = AGT
                c.alignment = Alignment(vertical="center")
            elif col == "Auto Error Count":
                # Dynamic COUNTIF formula — auto-updates when check cells are edited
                c.value = f'=COUNTIF({_chk_s}{ri}:{_chk_e}{ri},"❌*")'
                c.font      = Font(name=FNT, size=8, bold=True, color="9C0006")
                c.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment = Alignment(vertical="center")

    # Column widths
    widths = {
        "GRID":14, "Restaurant Name":32, "Lead Source":16,
        "Lead Status":12, "Lost Reason":18,
        "Edited By":22, "Edit Date":16, "Street":36, "Zip/Postal Code":12,
        "GM Title":28, "GM Category":22, "GM Status":18,
        "Auto Error Count":12, "Ops Feedback":24, "Updated Error Count":14,
        "Agent Acknowledgement":20, "Agent Feedback":28,
    }
    for ci, col in enumerate(ALL_COLS, 1):
        w = widths.get(col, 42 if col in CHECK_COLS else 14)
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A3"
    ws1.auto_filter.ref = f"A2:{get_column_letter(len(ALL_COLS))}{2+len(results_df)}"

    # ── Sheet 2: Agent Summary ─────────────────────────────────────
    ws2 = wb.create_sheet("Agent Summary")
    ws2["A1"] = "Agent KPI Summary"
    ws2["A1"].font = Font(name=FNT, bold=True, size=13, color="1F4E79")
    ws2.merge_cells("A1:D1")
    heads2 = list(agent_df.columns)
    for ci, h in enumerate(heads2, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.font = Font(name=FNT, bold=True, color="FFFFFF", size=9)
        c.fill = HDR; c.border = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, (_, r) in enumerate(agent_df.iterrows(), 3):
        for ci, col in enumerate(heads2, 1):
            c = ws2.cell(row=ri, column=ci, value=r[col])
            c.font = Font(name=FNT, size=8); c.border = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center")
    for ci, w in enumerate([28,14,16,16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Sampling Breakdown ────────────────────────────────
    ws3 = wb.create_sheet("Sampling Breakdown")
    ws3["A1"] = "Sampling Breakdown (10% per Agent × Source × Status)"
    ws3["A1"].font = Font(name=FNT, bold=True, size=13, color="1F4E79")
    ws3.merge_cells(f"A1:{get_column_letter(len(sampling_df.columns))}1")
    for ci, h in enumerate(sampling_df.columns, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.font = Font(name=FNT, bold=True, color="FFFFFF", size=9)
        c.fill = HDR; c.border = _thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
    for ri, (_, r) in enumerate(sampling_df.iterrows(), 3):
        for ci, col in enumerate(sampling_df.columns, 1):
            c = ws3.cell(row=ri, column=ci, value=r[col])
            c.font = Font(name=FNT, size=8); c.border = _thin()
            c.alignment = Alignment(horizontal="center", vertical="center")
    for ci, w in enumerate([24,18,14,12,10], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def main():
    if not check_password():
        return

    st.markdown("""
    <style>
    html,body,[class*="css"]{font-family:Arial,sans-serif;}
    header[data-testid="stHeader"]{background:#FFF;border-bottom:2px solid rgba(223,16,103,.25);}
    [data-testid="stSidebar"]{background:#FAFAFA;border-right:2px solid rgba(223,16,103,.3);}
    [data-testid="stSidebar"] h1,[data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3{color:#DF1067!important;font-weight:700!important;}
    [data-testid="stButton"]>button[kind="primary"]{
        background:#DF1067!important;border:none!important;color:#fff!important;
        font-weight:600!important;font-size:1rem!important;
        border-radius:8px!important;padding:.6rem 1.5rem!important;}
    [data-testid="stButton"]>button[kind="primary"]:hover{background:#C00055!important;}
    [data-testid="stDownloadButton"]>button{
        background:#DF1067!important;border:none!important;color:#fff!important;
        font-weight:600!important;border-radius:8px!important;width:100%!important;}
    [data-testid="stMetric"]{background:#FFF;border:1px solid #EBEBEB;
        border-top:3px solid #DF1067;border-radius:8px;
        padding:.9rem 1rem;box-shadow:0 1px 4px rgba(0,0,0,.06);}
    </style>""", unsafe_allow_html=True)

    if not _PYPINYIN_AVAILABLE:
        st.warning("⚠️ **pypinyin** not installed — Chinese name → Pinyin matching is off. "
                   "Run `pip install pypinyin` and restart.")

    # ── Sidebar ───────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(
            f'<div style="text-align:center;padding:.5rem 0 .8rem;">'
            f'<img src="data:image/png;base64,{DH_LOGO_B64}" style="width:110px;"/></div>',
            unsafe_allow_html=True)
        st.header("⚙️ Settings")

        market_code = st.selectbox(
            "Market",
            options=[None] + list(MARKETS.keys()),
            format_func=lambda k: (
                "— Select market —" if k is None
                else f"{MARKETS[k]['flag']} {MARKETS[k]['name']} ({k})"))
        market_cfg = MARKETS.get(market_code)

        st.divider()
        st.subheader("🎚 Match Thresholds")
        p2_threshold = st.slider("P3 Potential Match starts at", 0.30, 0.65, 0.50, 0.05,
                                  help="Name similarity at same postal ≥ this → P3 Potential Match")
        p3_threshold = st.slider("P4 Duplicate starts at",
                                  float(round(p2_threshold + 0.05, 2)), 0.95, 0.75, 0.05,
                                  help="Name similarity at same postal ≥ this → P4 Duplicate")
        st.caption(f"P3 threshold: {p2_threshold:.0%}  ·  P4 threshold: {p3_threshold:.0%}")

        st.divider()
        st.subheader("🚫 Exclusion Keywords")
        st.caption("Leads whose Google Maps category contains any of these → Wrong Target Group.")
        kw_input = st.text_area("Keywords (one per line)",
                                 value="\n".join(_DEFAULT_EXCLUSION_KW), height=180)
        exclusion_kw = [k.strip().lower() for k in kw_input.split("\n") if k.strip()]
        st.caption(f"{len(exclusion_kw)} keywords active")

        st.divider()
        with st.expander("🎯 Label reference"):
            st.markdown("""
| Label | What it means | Action |
|---|---|---|
| **P1 — New** | No CRM match. Google Maps confirms it's a restaurant. | ✅ Pitch delivery |
| **P3 — Potential Match** | Name 50–74% similar to a CRM account at same postal. | 🔍 Verify before pitching |
| **P4 — Duplicate** | Name ≥ 75% similar to a CRM account at same postal. | ❌ Skip — already in system |
| **Business Closed** | Google Maps shows permanently or temporarily closed. | ❌ Skip |
| **Wrong Target Group** | Google Maps category is not food delivery eligible. | ❌ Skip |
| **P2 — Please Check** | No Google Maps result found, or result is unreliable. | ⚪ Manual verification needed |
            """)

    if not market_cfg:
        st.info("👈 Select a market from the sidebar to get started.")
        return

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
        "📊 Classify Leads",
        "🔗 Generate Apify URLs",
        "🏢 SF Account Audit",
        "🔍 CRM Check",
        "📋 KPI Sample Checker",
        "📖 How to Use",
    ])

    # ════════════════════════════════════════════════════════════════
    # TAB 1 — LEAD CLASSIFICATION
    # ════════════════════════════════════════════════════════════════
    with tab1:
        with st.expander("📎 How to get your files — click to expand"):
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("**Step 1 · Leads file (Salesforce)**")
                st.link_button(
                    "Open Leads Report →",
                    "https://deliveryhero.lightning.force.com/"
                    "lightning/r/Report/00ObO000006oALhUAM/view",
                    use_container_width=True)
            with c2:
                st.markdown("**Step 2 · CRM Export (Salesforce)**")
                st.link_button(
                    "Open Singapore CRM Report →",
                    "https://deliveryhero.lightning.force.com/lightning/r/Report/"
                    "00ObO000005IE85UAG/view?queryScope=userFolders",
                    use_container_width=True)
            st.info(
                "🔗 **For Apify Results:** go to the **Generate Apify URLs** tab → "
                "Step 1 generates your URLs → paste into Apify → "
                "Step 2 adds the GRID column to your Apify export automatically.")

        st.divider()
        col1, col2, col3 = st.columns(3)
        with col1:
            st.subheader("1. Leads")
            leads_up = st.file_uploader("Upload leads (.xlsx or .csv)",
                                         type=["xlsx","xls","csv"], key="leads")
        with col2:
            st.subheader("2. Apify Results")
            apify_up = st.file_uploader("Upload Apify output (.csv or .xlsx)",
                                         type=["csv","xlsx","xls"], key="apify")
        with col3:
            st.subheader("3. CRM Export")
            crm_up   = st.file_uploader("Upload SF CRM export (.csv or .xlsx)",
                                         type=["csv","xlsx","xls"], key="crm")

        st.divider()

        import os as _os
        _builtin = _os.path.exists(_os.path.join(
            _os.path.dirname(_os.path.abspath(__file__)),
            f"zones_{market_code}.json"))

        if _builtin:
            st.markdown(f"**📍 Zones:** Built-in zone data for **{market_cfg['flag']} {market_cfg['name']}** loaded.")
            zone_up = st.file_uploader("Upload custom zone file to override (optional)",
                                        type=["csv","xlsx","xls"], key="zones")
        else:
            zone_up = st.file_uploader("Upload delivery zone file (.csv / .xlsx) — optional",
                                        type=["csv","xlsx","xls"], key="zones")
        geocode_on = st.toggle("Geocode leads without coordinates", value=True,
                                help="Uses OneMap SLA postal code API for sub-metre accuracy.")

        leads_df = col_map_leads = None
        crm_df   = col_map_crm   = None
        apify_df = col_map_apify = None

        if leads_up:
            try:
                leads_df, col_map_leads = load_leads(
                    leads_up.read(), leads_up.name, market_cfg)
                st.success(f"Leads loaded: {len(leads_df):,} rows")
                col_map_leads = _col_map_ui(
                    leads_df, col_map_leads,
                    fields=[
                        ("name",    "Restaurant name",    True),
                        ("street",  "Street / address",   False),
                        ("zip",     "Postal code",        False),
                        ("grid",    "GRID",               False),
                        ("lead_id", "Lead ID",            False),
                        ("lat",     "Latitude",           False),
                        ("lng",     "Longitude",          False),
                    ],
                    key_prefix="t1_leads")
            except Exception as e:
                st.error(f"Error loading leads: {e}")

        if crm_up:
            try:
                crm_df, col_map_crm = load_crm(
                    crm_up.read(), crm_up.name, market_cfg)
                st.success(f"CRM loaded: {len(crm_df):,} accounts")
            except Exception as e:
                st.error(f"Error loading CRM: {e}")

        if apify_up:
            try:
                apify_df, col_map_apify = load_apify(
                    apify_up.read(), apify_up.name)
                st.success(f"Apify loaded: {len(apify_df):,} rows")
            except Exception as e:
                st.error(f"Error loading Apify: {e}")

        st.divider()

        if leads_df is not None and st.button(
                "▶ Run Classification", type="primary", use_container_width=True):
            zones = []
            if zone_up:
                zones = load_zones(zone_up.read(), zone_up.name)
            elif _builtin:
                zones = load_zones(market_code=market_code)

            import time as _time
            _prog_bar  = st.progress(0.0)
            _prog_text = st.empty()
            _t0        = _time.time()

            def _progress_cb(done, total):
                pct     = done / total
                elapsed = _time.time() - _t0
                if done > 10:
                    eta_sec = int((elapsed / done) * (total - done))
                    eta_str = f"~{eta_sec//60}m {eta_sec%60:02d}s remaining"
                else:
                    eta_str = "estimating…"
                _prog_bar.progress(pct)
                _prog_text.caption(
                    f"Classifying {done:,} / {total:,} leads · {eta_str}")

            result_df = classify_leads(
                leads_df, col_map_leads,
                crm_df,   col_map_crm   or {},
                apify_df, col_map_apify or {},
                market_cfg,
                p2_threshold    = p2_threshold,
                p3_threshold    = p3_threshold,
                exclusion_kw    = exclusion_kw,
                zones           = zones,
                geocode_enabled = geocode_on,
                progress_cb     = _progress_cb,
            )
            _prog_bar.progress(1.0)
            _prog_text.caption(
                f"Done — {len(result_df):,} leads in "
                f"{int(_time.time()-_t0)}s")

            counts = result_df["Label"].value_counts()
            st.success(f"✅ Done — {len(result_df):,} leads classified.")

            # Metrics row
            m = st.columns(6)
            m[0].metric("✅ P1 New",           counts.get("P1 — New", 0))
            m[1].metric("🔴 P4 Duplicate",     counts.get("P4 — Duplicate", 0))
            m[2].metric("🟡 P3 Potential",     counts.get("P3 — Potential Match", 0))
            m[3].metric("🏢 Business Closed",  counts.get("Business Closed", 0))
            m[4].metric("❌ Wrong Target",     counts.get("Wrong Target Group", 0))
            m[5].metric("⚪ Please Check",     counts.get("P2 — Please Check", 0))

            # Per-priority tabs
            label_tabs = st.tabs([
                f"All ({len(result_df)})",
                f"✅ P1 New ({counts.get('P1 — New',0)})",
                f"🔴 P4 Duplicate ({counts.get('P4 — Duplicate',0)})",
                f"🟡 P3 Potential ({counts.get('P3 — Potential Match',0)})",
                f"🏢 Closed ({counts.get('Business Closed',0)})",
                f"❌ Wrong TG ({counts.get('Wrong Target Group',0)})",
                f"⚪ Please Check ({counts.get('P2 — Please Check',0)})",
            ])

            LABEL_STYLE = {
                "P1 — New":            "background-color:#d4edda",
                "P3 — Potential Match":"background-color:#fff3cd",
                "P4 — Duplicate":      "background-color:#f8d7da",
                "Business Closed":     "background-color:#ffeb9c",
                "Wrong Target Group":  "background-color:#ffdca8",
                "P2 — Please Check":   "background-color:#e2e3e5",
            }
            def _style(row):
                return [LABEL_STYLE.get(row["Label"],"")] * len(row)
            def _show(dff):
                if dff.empty: st.info("No leads in this category.")
                else: st.dataframe(dff.style.apply(_style, axis=1), use_container_width=True)

            with label_tabs[0]: _show(result_df)
            with label_tabs[1]: _show(result_df[result_df["Label"] == "P1 — New"])
            with label_tabs[2]: _show(result_df[result_df["Label"] == "P4 — Duplicate"])
            with label_tabs[3]: _show(result_df[result_df["Label"] == "P3 — Potential Match"])
            with label_tabs[4]: _show(result_df[result_df["Label"] == "Business Closed"])
            with label_tabs[5]: _show(result_df[result_df["Label"] == "Wrong Target Group"])
            with label_tabs[6]: _show(result_df[result_df["Label"] == "P2 — Please Check"])

            excel_buf = build_excel(result_df, f"{market_cfg['flag']} {market_cfg['name']}")
            st.download_button(
                "⬇ Download Excel Report",
                data=excel_buf,
                file_name=f"ALG_{pd.Timestamp.now().strftime('%d%m%y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)

    # ════════════════════════════════════════════════════════════════
    # TAB 2 — GENERATE APIFY URLS
    # ════════════════════════════════════════════════════════════════
    with tab2:
        st.subheader("Generate Google Maps URLs for Apify")

        # ── Step 1: Generate URLs ─────────────────────────────────
        st.markdown("#### Step 1 · Generate URLs")

        url_mode = st.radio(
            "URL format",
            options=["text", "coords"],
            format_func=lambda m: (
                "📝  Company / Account + Street + Postal"
                if m == "text"
                else "📍  Company / Account + Coordinates (Latitude, Longitude)"
            ),
            horizontal=False,
            key="url_mode",
        )

        url_up = st.file_uploader("Upload leads file (.xlsx or .csv)",
                                   type=["xlsx","xls","csv"], key="url_leads")
        if url_up and market_cfg:
            try:
                url_bytes = url_up.read()
                url_df, url_col_map = load_leads(url_bytes, url_up.name, market_cfg)
                urls, reused = generate_google_urls(url_df, url_col_map, market_cfg,
                                                    mode=url_mode)
                url_df["GOOGLE URL"] = urls
                valid = [u for u in urls if u]

                # Store GRID → norm_url mapping in session state for Step 2
                grid_col_name = url_col_map.get("grid")
                if grid_col_name:
                    url_to_grid = {}
                    for i, (_, row) in enumerate(url_df.iterrows()):
                        if i < len(urls) and urls[i]:
                            g = str(row.get(grid_col_name, "") or "").strip()
                            if g:
                                url_to_grid[norm_url(urls[i])] = g
                    st.session_state["url_to_grid"] = url_to_grid

                st.info(f"{len(valid):,} URLs generated ({reused:,} reused from existing column).")
                st.text_area("Generated targets", "\n".join(valid), height=180)
                buf = io.StringIO()
                url_df.to_csv(buf, index=False)
                st.download_button("⬇ Download leads with URLs (.csv)",
                                   buf.getvalue(),
                                   f"leads_with_urls_{market_code}.csv",
                                   mime="text/csv", use_container_width=True)
            except Exception as e:
                st.error(f"Error: {e}")

        st.divider()

        # ── Step 2: Append GRID to Apify export ──────────────────
        st.markdown("#### Step 2 · Add GRID to your Apify Export")
        st.caption(
            "After running Apify, upload your export here. "
            "The tool matches each row via `inputStartUrl` and adds a **GRID** column.")

        # Build lookup — prefer session state, fall back to manual URL CSV upload
        lookup: dict = dict(st.session_state.get("url_to_grid", {}))

        if lookup:
            st.success(f"✅ {len(lookup):,} GRIDs ready from Step 1 above.")
        else:
            st.info("No URLs generated this session yet. Upload your URL CSV below.")

        with st.expander("📂 Upload URL CSV (if you generated URLs in a previous session)"):
            st.caption(
                "Upload the CSV downloaded from Step 1, or your Google Sheet URL extractor. "
                "The tool auto-detects the header row and the GRID + URL columns.")
            url_csv_up = st.file_uploader(
                "Upload leads+URL CSV (.csv or .xlsx)",
                type=["csv","xlsx","xls"], key="url_csv_fallback")
            if url_csv_up:
                try:
                    fb_bytes = url_csv_up.read()
                    from io import StringIO as _SIO

                    if url_csv_up.name.endswith((".xlsx", ".xls")):
                        # Excel / HTML-as-XLS — _cached_read handles both formats
                        fb_df = _cached_read(fb_bytes, url_csv_up.name)
                        # Google Sheets exported as XLS often have title rows
                        # before the real column headers. If "grid" and "url"
                        # aren't in the columns yet, scan the rows for them.
                        _has_grid = detect_column(fb_df, ["GRID","grid","Grid"])
                        _has_url  = detect_column(fb_df, [
                            "GOOGLE URL","Google URL","google_url",
                            "URL (Name + Address + Postal)",
                            "URL (Name + Lat + Long)",
                            "URL (Name + Coordinates)","URL","url"])
                        if not (_has_grid and _has_url):
                            for i, row in fb_df.iterrows():
                                vals = [str(v).lower() for v in row.tolist()]
                                if (any("grid" in v for v in vals)
                                        and any("url" in v for v in vals)):
                                    fb_df.columns = [str(c) for c in fb_df.iloc[i].tolist()]
                                    fb_df = fb_df.iloc[i+1:].reset_index(drop=True)
                                    break
                    else:
                        # ── Smart header detection for CSV ────────────────
                        # Google Sheet exports have title rows before the real headers.
                        # Scan every row to find one that contains both "grid" and "url".
                        raw_rows = None
                        for enc in ("utf-8","utf-8-sig","windows-1252","latin-1"):
                            try:
                                raw_rows = pd.read_csv(
                                    _SIO(fb_bytes.decode(enc)),
                                    header=None, on_bad_lines="skip", engine="python")
                                break
                            except Exception:
                                continue

                        header_row = 0
                        if raw_rows is not None:
                            for i, row in raw_rows.iterrows():
                                vals = [str(v).lower() for v in row.tolist()]
                                if any("grid" in v for v in vals) and any("url" in v for v in vals):
                                    header_row = i
                                    break

                        fb_df = None
                        for enc in ("utf-8","utf-8-sig","windows-1252","latin-1"):
                            try:
                                fb_df = pd.read_csv(
                                    _SIO(fb_bytes.decode(enc)),
                                    header=header_row, on_bad_lines="skip", engine="python")
                                break
                            except Exception:
                                continue
                        if fb_df is None:
                            fb_df = _cached_read(fb_bytes, url_csv_up.name)

                    # ── Column detection ───────────────────────────────
                    # Covers both tool-generated CSV and Google Sheet formats
                    grid_c = detect_column(fb_df, ["GRID","grid","Grid"])
                    gurl_c = detect_column(fb_df, [
                        "GOOGLE URL","Google URL","google_url",
                        "URL (Name + Address + Postal)",
                        "URL (Name + Lat + Long)",
                        "URL (Name + Coordinates)",
                        "URL","url",
                    ])

                    if grid_c and gurl_c:
                        for _, row in fb_df.iterrows():
                            g = str(row.get(grid_c,"") or "").strip()
                            u = str(row.get(gurl_c,"") or "").strip()
                            if g and u and u.lower() not in ("nan",""):
                                lookup[norm_url(u)] = g
                        st.success(f"Loaded {len(lookup):,} GRIDs from uploaded file.")
                    elif grid_c and not gurl_c:
                        # URL column name not recognised — find any column with http URLs
                        auto_url_col = next(
                            (col for col in fb_df.columns
                             if fb_df[col].dropna().astype(str).str.startswith("http").any()),
                            None)
                        if auto_url_col:
                            for _, row in fb_df.iterrows():
                                g = str(row.get(grid_c,"") or "").strip()
                                u = str(row.get(auto_url_col,"") or "").strip()
                                if g and u and u.startswith("http"):
                                    lookup[norm_url(u)] = g
                            st.success(
                                f"Loaded {len(lookup):,} GRIDs "
                                f"(URL column auto-detected as '{auto_url_col}').")
                        else:
                            st.error("Found GRID column but could not identify the URL column.")
                    else:
                        st.error(
                            f"Could not find GRID or URL columns. "
                            f"Detected columns: {list(fb_df.columns[:8])}")
                except Exception as e:
                    st.error(f"Error reading URL CSV: {e}")

        apify_merge_up = st.file_uploader(
            "Upload Apify export (.csv or .xlsx)",
            type=["csv","xlsx","xls"], key="apify_merge")

        if apify_merge_up and lookup:
            try:
                merge_bytes = apify_merge_up.read()
                apy_df      = _cached_read(merge_bytes, apify_merge_up.name)

                # Find the input URL column — inputStartUrl is the cleanest match
                input_url_col = detect_column(apy_df, [
                    "inputStartUrl", "input_start_url",
                    "searchPageUrl",  "search_page_url",
                    "searchUrl",      "inputUrl",
                ])

                if input_url_col is None:
                    st.error(
                        "Could not find `inputStartUrl` (or `searchPageUrl`) column "
                        "in the Apify export. Make sure you export with these columns enabled.")
                else:
                    grids   = []
                    matched = 0
                    for _, row in apy_df.iterrows():
                        raw     = str(row.get(input_url_col, "") or "")
                        key     = norm_url(raw)
                        grid    = lookup.get(key, "")
                        if grid:
                            matched += 1
                        grids.append(grid)

                    result_df = apy_df.copy()
                    if "GRID" in result_df.columns:
                        result_df["GRID"] = grids   # overwrite existing
                    else:
                        result_df.insert(0, "GRID", grids)

                    unmatched = len(result_df) - matched
                    if unmatched > 0:
                        st.warning(
                            f"⚠️ {matched} of {len(result_df)} rows matched. "
                            f"{unmatched} rows have no GRID — check that this Apify export "
                            f"was generated from the same URL batch.")
                    else:
                        st.success(f"✅ All {len(result_df)} rows matched successfully.")

                    buf2 = io.StringIO()
                    result_df.to_csv(buf2, index=False)
                    st.download_button(
                        "⬇ Download Apify export with GRID",
                        buf2.getvalue(),
                        "apify_with_grid.csv",
                        mime="text/csv",
                        use_container_width=True)

            except Exception as e:
                st.error(f"Error processing Apify export: {e}")

        elif apify_merge_up and not lookup:
            st.warning("No GRID lookup available. Generate URLs in Step 1 first, "
                       "or upload your URL CSV using the expander above.")

    # ════════════════════════════════════════════════════════════════
    # TAB 3 — SF ACCOUNT AUDIT
    # ════════════════════════════════════════════════════════════════
    with tab3:
        st.markdown(
            "Find suspected duplicate accounts **within your Salesforce master**. "
            "Run this periodically for data hygiene — it's separate from lead dedup.")

        audit_up = st.file_uploader(
            "Upload Salesforce Master (for audit)", type=["csv","xlsx","xls"], key="audit")

        if audit_up:
            try:
                df_audit_raw = _cached_read(audit_up.read(), audit_up.name)
                st.success(f"Loaded: {len(df_audit_raw):,} accounts")
            except Exception as e:
                st.error(f"Error: {e}"); st.stop()

            # Auto-detect then let user confirm / override
            auto_audit = {
                "name":   detect_column(df_audit_raw, ["Account Name","Name","name"]),
                "postal": detect_column(df_audit_raw, ["Restaurant PostalCode","PostalCode",
                                                        "Postal Code","Zip/Postal Code",
                                                        "BillingPostalCode"]),
                "street": detect_column(df_audit_raw, ["Formatted Restaurant Address",
                                                        "BillingStreet","Street","Address"]),
                "status": detect_column(df_audit_raw, ["Account Status","Account_Status__c"]),
                "grid":   detect_column(df_audit_raw, ["GRID","grid","Grid"]),
                "sf_id":  detect_column(df_audit_raw, ["SF 18 Char ID","Id","SF_ID",
                                                        "Salesforce ID"]),
            }
            audit_map = _col_map_ui(
                df_audit_raw, auto_audit,
                fields=[
                    ("name",   "Account Name",    True),
                    ("postal", "Postal Code",     True),
                    ("street", "Address",         False),
                    ("status", "Account Status",  False),
                    ("grid",   "GRID",            False),
                    ("sf_id",  "SF 18-Char ID",   False),
                ],
                key_prefix="t3_audit")
            au_name   = audit_map.get("name")
            au_post   = audit_map.get("postal")
            au_addr   = audit_map.get("street")
            au_status = audit_map.get("status")
            au_grid   = audit_map.get("grid")
            au_id     = audit_map.get("sf_id")

            audit_thresh = st.slider(
                "Name similarity threshold", 50, 100, 70, 5, key="audit_thresh",
                help="Pairs at or above this similarity at same postal+unit are flagged.")
            st.caption(
                f"Pairs with name similarity ≥ {audit_thresh}% at the same postal code "
                f"and unit number will be flagged as suspected duplicates.")

            if st.button("🔍 Run SF Account Audit", type="primary", key="run_audit"):
                with st.spinner("Preprocessing SF accounts…"):
                    df_audit = _cached_preprocess_crm(
                        df_audit_raw,
                        name_col   = au_name,
                        postal_col = au_post,
                        addr_col   = au_addr,
                        status_col = au_status,
                        char_map_tuple = (),
                    )

                progress_a = st.progress(0, text="Scanning for duplicates…")
                with st.spinner("Scanning for suspected duplicates…"):
                    pairs = find_sf_duplicates(
                        df_audit, au_name, au_addr,
                        au_status, au_grid, au_id,
                        audit_thresh,
                    )
                progress_a.progress(1.0, text="Done!")

                if not pairs:
                    st.success(
                        f"✅ No suspected duplicates found at {audit_thresh}% threshold. "
                        f"Your SF data looks clean.")
                else:
                    pairs_df = pd.DataFrame(pairs)
                    high   = len(pairs_df[pairs_df["RISK_LEVEL"] == "🔴 High"])
                    medium = len(pairs_df[pairs_df["RISK_LEVEL"] == "🟡 Medium"])
                    low    = len(pairs_df[pairs_df["RISK_LEVEL"] == "🟢 Low"])

                    st.success(f"✅ Audit complete — {len(pairs_df):,} suspected duplicate pairs found.")

                    m1, m2, m3, m4 = st.columns(4)
                    m1.metric("Total Pairs",    len(pairs_df))
                    m2.metric("🔴 High Risk",   high)
                    m3.metric("🟡 Medium Risk", medium)
                    m4.metric("🟢 Low Risk",    low)

                    # Risk breakdown chart
                    st.subheader("Risk Breakdown")
                    st.bar_chart(pd.DataFrame({
                        "Risk": ["🔴 High","🟡 Medium","🟢 Low"],
                        "Pairs": [high, medium, low]
                    }).set_index("Risk"))

                    def _audit_style(row):
                        c = {"🔴 High":"background-color:#f8d7da",
                             "🟡 Medium":"background-color:#fff3cd",
                             "🟢 Low":"background-color:#d4edda"}
                        return [c.get(row["RISK_LEVEL"],"")] * len(row)
                    def _show_a(dff):
                        if dff.empty: st.info("No pairs in this category.")
                        else: st.dataframe(dff.style.apply(_audit_style, axis=1),
                                           use_container_width=True)

                    st.subheader("Suspected Duplicate Pairs")
                    a_tabs = st.tabs([
                        f"All ({len(pairs_df)})",
                        f"🔴 High ({high})",
                        f"🟡 Medium ({medium})",
                        f"🟢 Low ({low})",
                    ])
                    with a_tabs[0]: _show_a(pairs_df)
                    with a_tabs[1]: _show_a(pairs_df[pairs_df["RISK_LEVEL"] == "🔴 High"])
                    with a_tabs[2]: _show_a(pairs_df[pairs_df["RISK_LEVEL"] == "🟡 Medium"])
                    with a_tabs[3]: _show_a(pairs_df[pairs_df["RISK_LEVEL"] == "🟢 Low"])

                    st.download_button(
                        "📥 Download Audit Results",
                        pairs_df.to_csv(index=False),
                        "sf_audit_results.csv",
                        mime="text/csv", use_container_width=True)

    # ════════════════════════════════════════════════════════════════
    # TAB 4 — CRM CHECK (no GRID / no Apify)
    # ════════════════════════════════════════════════════════════════
    with tab4:  # CRM Check
        st.subheader("🔍 Quick CRM Duplicate Check")
        st.caption(
            "Upload a restaurant list (e.g. from a government website) and your "
            "Salesforce CRM export to check for existing duplicates. "
            "No GRID or Apify needed.")

        cc1, cc2 = st.columns(2)
        with cc1:
            st.markdown("**Restaurant List**")
            st.caption("CSV or Excel — needs: Name, Street, Postal Code")
            rest_up = st.file_uploader(
                "Upload restaurant list",
                type=["csv","xlsx","xls"], key="crm_check_rest")
        with cc2:
            st.markdown("**CRM All Accounts (Salesforce)**")
            st.caption("Same export used in the Classify Leads tab")
            crm_chk_up = st.file_uploader(
                "Upload CRM export",
                type=["csv","xlsx","xls"], key="crm_check_crm")

        if rest_up and crm_chk_up:
            try:
                rest_bytes = rest_up.read()
                rest_raw   = _cached_read(rest_bytes, rest_up.name)

                crm_chk_bytes = crm_chk_up.read()
                crm_chk_df, col_map_crm_chk = load_crm(crm_chk_bytes, crm_chk_up.name,
                                                         market_cfg)

                # ── Auto-detect restaurant list columns ───────────
                auto_rest = {
                    "name":   detect_column(rest_raw, [
                        "Company / Account","Account Name","Restaurant Name",
                        "Name","Business Name","name"]),
                    "street": detect_column(rest_raw, [
                        "Street","Formatted Restaurant Address","Address",
                        "Full Address","Formatted Address","street","address"]),
                    "postal": detect_column(rest_raw, [
                        "Zip/Postal Code","Restaurant PostalCode","Postal Code",
                        "Postal","PostalCode","Zip","postal","zip"]),
                    "grid":   detect_column(rest_raw, ["GRID","grid","Grid"]),
                }
                rest_map = _col_map_ui(
                    rest_raw, auto_rest,
                    fields=[
                        ("name",   "Restaurant name",  True),
                        ("street", "Street / address", False),
                        ("postal", "Postal code",      False),
                        ("grid",   "GRID",             False),
                    ],
                    key_prefix="t5_rest")
                name_c   = rest_map.get("name")
                street_c = rest_map.get("street")
                postal_c = rest_map.get("postal")
                grid_c   = rest_map.get("grid")
                rest_cols = {"name": name_c, "street": street_c,
                             "postal": postal_c, "grid": grid_c}

                if not name_c:
                    st.warning("Please map the restaurant name column above.")
                else:
                    st.info(
                        f"**{len(rest_raw):,} restaurants** loaded  ·  "
                        f"Name: `{name_c}`  ·  "
                        f"Street: `{street_c or '—'}`  ·  "
                        f"Postal: `{postal_c or '—'}`")

                    if st.button("▶ Run CRM Check", type="primary",
                                 use_container_width=True, key="run_crm_check"):
                        with st.spinner("Checking against CRM…"):
                            _char_map = market_cfg.get("char_map", {})
                            check_df = crm_check_classify(
                                rest_raw, rest_cols,
                                crm_chk_df, col_map_crm_chk,
                                _char_map, p2_threshold, p3_threshold)
                            st.session_state["crm_check_result"] = check_df

                    if "crm_check_result" in st.session_state:
                        chk = st.session_state["crm_check_result"]
                        cnts = chk["Label"].value_counts()

                        mc = st.columns(3)
                        mc[0].metric("✅ Unverified — Create",
                                     cnts.get("Unverified", 0))
                        mc[1].metric("🟡 P3 — Review first",
                                     cnts.get("P3 — Potential Match", 0))
                        mc[2].metric("🔴 P4 — Duplicate Skip",
                                     cnts.get("P4 — Duplicate", 0))

                        CSTYLE = {
                            "Unverified":           "background-color:#dbeafe",
                            "P3 — Potential Match": "background-color:#fff3cd",
                            "P4 — Duplicate":       "background-color:#f8d7da",
                        }
                        def _cs(row):
                            return [CSTYLE.get(row["Label"],"")] * len(row)
                        def _cshow(dff):
                            if dff.empty: st.info("No entries in this category.")
                            else: st.dataframe(
                                dff.style.apply(_cs, axis=1),
                                use_container_width=True)

                        ctabs = st.tabs([
                            f"All ({len(chk)})",
                            f"✅ Unverified ({cnts.get('Unverified',0)})",
                            f"🟡 P3 Potential ({cnts.get('P3 — Potential Match',0)})",
                            f"🔴 P4 Duplicate ({cnts.get('P4 — Duplicate',0)})",
                        ])
                        with ctabs[0]: _cshow(chk)
                        with ctabs[1]: _cshow(chk[chk["Label"]=="Unverified"])
                        with ctabs[2]: _cshow(chk[chk["Label"]=="P3 — Potential Match"])
                        with ctabs[3]: _cshow(chk[chk["Label"]=="P4 — Duplicate"])

                        excel_chk = build_crm_check_excel(chk)
                        st.download_button(
                            "📥 Download CRM Check Report",
                            excel_chk,
                            f"CRM_Check_{pd.Timestamp.now().strftime('%d%m%y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument"
                                 ".spreadsheetml.sheet",
                            use_container_width=True)

            except Exception as e:
                st.error(f"Error: {e}")

    # ════════════════════════════════════════════════════════════════
    # TAB 5 — KPI SAMPLE CHECKER
    # ════════════════════════════════════════════════════════════════
    with tab5:
        st.subheader("📋 KPI Sample Checker")
        st.caption("Monthly QA check — 10% stratified sample per Agent × Source × Status.")

        st.markdown("**Step 1 · Upload Lead Status Change report**")
        st.link_button(
            "Open Lead Status Change Report →",
            "https://deliveryhero.lightning.force.com/lightning/r/Report/"
            "00ObO000007f12XUAQ/view?queryScope=userFolders",
            use_container_width=False)
        kpi_up = st.file_uploader(
            "Upload Lead Status Change report (.xlsx or .csv)",
            type=["xlsx","xls","csv"], key="kpi_leads")

        kpi_sampled_df = None
        kpi_sampling_df = None

        if kpi_up:
            try:
                kpi_raw = _cached_read(kpi_up.read(), kpi_up.name)
                kpi_raw = clean_lead_report(kpi_raw)
                st.success(f"Loaded: {len(kpi_raw):,} leads")
                kpi_sampled_df, kpi_sampling_df = sample_leads(kpi_raw)
                st.success(f"Sample: {len(kpi_sampled_df):,} leads selected "
                           f"({len(kpi_sampling_df):,} strata)")
                with st.expander("📊 Sampling breakdown"):
                    st.dataframe(kpi_sampling_df, use_container_width=True)

                # Generate Apify URLs for sampled leads
                st.divider()
                st.markdown("**Step 2 · Generate Google Maps URLs**")
                url_mode_kpi = st.radio(
                    "URL format", ["text","coords"],
                    format_func=lambda m: (
                        "📝 Street + Postal" if m == "text"
                        else "📍 Coordinates (if available)"),
                    horizontal=True, key="kpi_url_mode")

                kpi_urls = []
                for _, row in kpi_sampled_df.iterrows():
                    # Include restaurant name — strip garbled/non-ASCII chars
                    # (Chinese names may read as ??? from XLS encoding)
                    raw_name = str(row.get("Company","") or
                                   row.get("Company / Account","") or "")
                    name    = re.sub(r'[^\x20-\x7E]+', '', raw_name).strip()
                    street  = str(row.get("Street","") or "")
                    postal  = _norm_postal_input(
                                  str(row.get("Zip/Postal Code","") or ""))
                    if url_mode_kpi == "text":
                        parts = [p for p in [name, street, postal]
                                 if p and p.lower() not in ("0","0.0","nan","")]
                        q = " ".join(parts)
                        url = (f"https://www.google.com/maps/search/?api=1&query={quote(q)}"
                               if q else "")
                    else:
                        url = ""
                    kpi_urls.append(url)

                kpi_sampled_df = kpi_sampled_df.copy()
                kpi_sampled_df["GOOGLE URL"] = kpi_urls

                # Store GRID→norm_url mapping in session state
                url_to_grid = {}
                for (_, r), u in zip(kpi_sampled_df.iterrows(), kpi_urls):
                    if not u:
                        continue
                    # Safely extract GRID as a plain string
                    g = r["GRID"] if "GRID" in r.index else ""
                    if isinstance(g, pd.Series):
                        g = g.iloc[0] if len(g) > 0 else ""
                    g = str(g or "").strip()
                    if g:
                        url_to_grid[norm_url(u)] = g
                st.session_state["kpi_url_to_grid"] = url_to_grid
                st.caption(f"✅ {len(url_to_grid):,} URL→GRID pairs stored in session")

                valid_urls = [u for u in kpi_urls if u]
                st.text_area("URLs for Apify", "\n".join(valid_urls), height=150,
                             key="kpi_url_preview")
                buf_kpi = io.StringIO()
                kpi_sampled_df.to_csv(buf_kpi, index=False)
                st.download_button(
                    "⬇ Download sampled leads + URLs",
                    buf_kpi.getvalue(),
                    f"kpi_sample_{pd.Timestamp.now().strftime('%d%m%y')}.csv",
                    mime="text/csv", use_container_width=True)

            except Exception as e:
                st.error(f"Error: {e}")

        st.divider()
        st.markdown("**Step 3 · Run Apify**")
        st.info(
            "1. Copy the URLs from Step 2 and paste them as inputs into your Apify scraper\n"
            "2. Run the scraper — takes ~10–15 min for 100–200 leads\n"
            "3. Export the Apify results as CSV\n"
            "4. Come back here and upload the raw export in Step 4 below — "
            "the tool will automatically match each row to a GRID via inputStartUrl"
        )

        st.divider()
        st.markdown("**Step 4 · Upload supporting files & run checks**")

        k1, k2, k3 = st.columns(3)
        with k1:
            kpi_apify_up = st.file_uploader(
                "Apify export — raw export from Apify (optional)",
                type=["csv","xlsx","xls"], key="kpi_apify")
        with k2:
            st.link_button(
                "Open CRM All Accounts →",
                "https://deliveryhero.lightning.force.com/lightning/r/Report/"
                "00ObO000005IE85UAG/view?queryScope=userFolders",
                use_container_width=True)
            kpi_crm_up = st.file_uploader(
                "CRM All Accounts (required)",
                type=["csv","xlsx","xls"], key="kpi_crm")
        with k3:
            st.link_button(
                "Open Account Details Report →",
                "https://deliveryhero.lightning.force.com/lightning/r/Report/"
                "00ObO000007f1TxUAI/view",
                use_container_width=True)
            kpi_acc_up = st.file_uploader(
                "Account Details — converted leads (required)",
                type=["csv","xlsx","xls"], key="kpi_acc")

        if kpi_crm_up and kpi_acc_up and kpi_sampled_df is not None:
            # ── Apify: auto-match GRID via session state URL→GRID map ──
            kpi_apify_df = None
            if kpi_apify_up:
                try:
                    kpi_apify_raw = _cached_read(kpi_apify_up.read(), kpi_apify_up.name)

                    # Check if GRID already present
                    existing_gc = detect_column(kpi_apify_raw, ["GRID","grid","Grid"])

                    if existing_gc and existing_gc == "GRID":
                        # Already has GRID — use directly
                        kpi_apify_df = kpi_apify_raw
                        st.success(f"Apify loaded: {len(kpi_apify_df):,} rows (GRID column present)")
                    else:
                        # Match via inputStartUrl → GRID using session state
                        url_map = st.session_state.get("kpi_url_to_grid", {})
                        url_col = detect_column(kpi_apify_raw,
                                                ["inputStartUrl","searchPageUrl"])
                        if url_col and url_map:
                            grids   = []
                            for _, r in kpi_apify_raw.iterrows():
                                key = norm_url(str(r.get(url_col,"") or ""))
                                grids.append(url_map.get(key,""))
                            matched = sum(1 for g in grids if g)
                            if matched == 0:
                                st.error(
                                    "❌ 0 GRIDs matched — the Apify file you uploaded "
                                    "appears to be from a **different batch**.\n\n"
                                    "**What to do:**\n"
                                    "1. Go back to Step 2 and download the URL CSV\n"
                                    "2. Paste those specific URLs into Apify\n"
                                    "3. Export and upload **that** Apify result here")
                                kpi_apify_df = None
                            else:
                                if existing_gc:
                                    kpi_apify_raw["GRID"] = grids
                                else:
                                    kpi_apify_raw.insert(0, "GRID", grids)
                                kpi_apify_df = kpi_apify_raw
                                st.success(
                                    f"Apify loaded: {len(kpi_apify_df):,} rows — "
                                    f"{matched:,} GRIDs matched via URL")
                                if matched < len(kpi_apify_df):
                                    st.warning(
                                        f"{len(kpi_apify_df)-matched:,} rows could not be matched.")
                            # Debug expander — helps diagnose mismatches
                            with st.expander("🔍 URL matching debug (expand if GRIDs not matching)"):
                                st.write(f"Session state URL map: **{len(url_map):,} keys**")
                                if url_map:
                                    sample_key = next(iter(url_map))
                                    st.write(f"Sample key in map: `{sample_key[:80]}...`")
                                    st.write(f"→ GRID: `{url_map[sample_key]}`")
                                first_apify_url = norm_url(
                                    str(kpi_apify_raw.iloc[0].get(url_col,"") or ""))
                                st.write(f"First Apify inputStartUrl (normalised): `{first_apify_url[:80]}...`")
                                st.write(f"Key exists in map: **{first_apify_url in url_map}**")
                        else:
                            # Fallback: use as-is and warn
                            kpi_apify_df = kpi_apify_raw
                            st.warning(
                                "No URL→GRID map found in session state. "
                                "Generate URLs in Step 1 first, then upload Apify results here. "
                                "Apify data loaded but GRID matching may be incomplete.")
                except Exception as e:
                    st.error(f"Apify error: {e}")

            # Load CRM
            kpi_crm_raw     = None
            kpi_col_map_crm = {}
            try:
                kpi_crm_raw  = _cached_read(kpi_crm_up.read(), kpi_crm_up.name)
                auto_crm = {
                    "name":   detect_column(kpi_crm_raw, ["Account Name","Name","name"]),
                    "postal": detect_column(kpi_crm_raw, ["Restaurant PostalCode","PostalCode",
                                                           "Postal Code","Zip/Postal Code"]),
                    "street": detect_column(kpi_crm_raw, ["Formatted Restaurant Address",
                                                           "BillingStreet","Street","Address"]),
                    "grid":   detect_column(kpi_crm_raw, ["GRID","grid","Grid"]),
                    "status": detect_column(kpi_crm_raw, ["Account Status","Status"]),
                }
                kpi_col_map_crm = _col_map_ui(
                    kpi_crm_raw, auto_crm,
                    fields=[
                        ("name",   "Account Name",   True),
                        ("postal", "Postal Code",    True),
                        ("street", "Address",        False),
                        ("grid",   "GRID",           False),
                        ("status", "Account Status", False),
                    ], key_prefix="kpi_crm")
                st.success(f"CRM loaded: {len(kpi_crm_raw):,} accounts")
            except Exception as e:
                st.error(f"CRM error: {e}")

            # Load account details
            kpi_acc_df   = None
            kpi_acc_cols = {}
            if kpi_acc_up:
                try:
                    kpi_acc_df  = _cached_read(kpi_acc_up.read(), kpi_acc_up.name)
                    auto_acc = {
                        "grid":            detect_column(kpi_acc_df, ["GRID","grid"]),
                        "name":            detect_column(kpi_acc_df, ["Account Name","Name"]),
                        "phone":           detect_column(kpi_acc_df, ["Phone"]),
                        "email":           detect_column(kpi_acc_df, ["Account Email","Email"]),
                        "website":         detect_column(kpi_acc_df, ["Website"]),
                        "social_media":    detect_column(kpi_acc_df, ["Social Media URL","Social Media"]),
                        "parent_account":  detect_column(kpi_acc_df, ["Parent Account"]),
                        "business_office": detect_column(kpi_acc_df, ["Business Office"]),
                        "delivery_service":detect_column(kpi_acc_df, ["Delivery Service","Category"]),
                        "target_partner":  detect_column(kpi_acc_df, ["Target Partner"]),
                        "category":        detect_column(kpi_acc_df, ["Category","Restaurant Category"]),
                    }
                    kpi_acc_cols = _col_map_ui(
                        kpi_acc_df, auto_acc,
                        fields=[
                            ("grid",            "GRID",             True),
                            ("name",            "Account Name",     True),
                            ("phone",           "Phone",            False),
                            ("email",           "Email",            False),
                            ("website",         "Website",          False),
                            ("social_media",    "Social Media URL", False),
                            ("parent_account",  "Parent Account",   False),
                            ("business_office", "Business Office",  False),
                            ("delivery_service","Delivery Service", False),
                            ("target_partner",  "Target Partner",   False),
                            ("category",        "Restaurant Category",False),
                        ], key_prefix="kpi_acc")
                    st.success(f"Account details loaded: {len(kpi_acc_df):,} rows")
                except Exception as e:
                    st.error(f"Account details error: {e}")

            if kpi_crm_raw is not None and st.button(
                    "▶ Run KPI Checks", type="primary",
                    use_container_width=True, key="run_kpi"):
                with st.spinner("Running KPI checks…"):
                    try:
                        kpi_char_map = market_cfg.get("char_map", {})
                        kpi_zones    = []
                        if _builtin:
                            kpi_zones = load_zones(market_code=market_code)

                        kpi_results_df, kpi_agent_df = run_kpi_checks(
                            sampled_df   = kpi_sampled_df,
                            apify_df     = kpi_apify_df,
                            crm_df       = kpi_crm_raw,
                            col_map_crm  = kpi_col_map_crm,
                            account_df   = kpi_acc_df,
                            account_cols = kpi_acc_cols,
                            zones        = kpi_zones,
                            char_map     = kpi_char_map,
                            p2_threshold = p2_threshold,
                            p3_threshold = p3_threshold,
                        )
                        st.success(
                            f"Done — {len(kpi_results_df):,} leads checked, "
                            f"{kpi_results_df['Auto Error Count'].sum():,} auto errors found")
                        st.dataframe(kpi_agent_df, use_container_width=True)
                        kpi_excel = build_kpi_excel(
                            kpi_results_df, kpi_agent_df, kpi_sampling_df)
                        st.download_button(
                            "📥 Download KPI Scorecard (.xlsx)",
                            kpi_excel,
                            f"KPI_{pd.Timestamp.now().strftime('%d%m%y')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True)
                    except Exception as e:
                        st.error(f"KPI check error: {e}")
        elif kpi_sampled_df is None:
            st.info("Upload the Lead Status Change report above to begin.")
        else:
            st.info("Upload CRM All Accounts and Account Details report to run checks.")

    # ════════════════════════════════════════════════════════════════
    # TAB 6 — HOW TO USE
    # ════════════════════════════════════════════════════════════════
    with tab6:  # How to Use
        st.markdown("""
## 📖 How to Use the Lead Classifier

This tool classifies new restaurant leads against your Salesforce CRM to identify
duplicates, potential matches, and genuinely new business opportunities.

---

### Step 1 · Export your Leads from Salesforce
1. Click **Open Leads Report →** in the *Classify Leads* tab
2. Export the report as **CSV** or **Excel (.xls / .xlsx)**
3. Required columns: `GRID`, `Company / Account`, `Street`, `Zip/Postal Code`

---

### Step 2 · Generate URLs, Run Apify, Get GRID back
Go to the **🔗 Generate Apify URLs** tab:

**Step 2A — Generate URLs**
1. Upload your leads file
2. Choose your URL format:
   - **📝 Company / Account + Street + Postal** — recommended
   - **📍 Company / Account + Coordinates** — use if leads have lat/lng
3. Copy the URLs from the *Generated targets* box and paste directly into Apify

**Step 2B — Add GRID to your Apify Export**
1. After Apify finishes (10–15 min), export the results as CSV
2. Come back to **Step 2** in the Generate Apify URLs tab
3. Upload the Apify export — the tool matches each row to a GRID via `inputStartUrl`
4. Download the enriched Apify file with GRID as the first column

> 💡 Keep the Streamlit tab open while Apify runs — the GRID lookup stays in memory.
> If you close the browser, use the *Upload URL CSV* expander in Step 2B to reload.

> ⚠️ Works with both URL formats and mixed batches — matching is always URL-based, never row-order.

---

### Step 3 · Export Salesforce All Accounts (CRM)
1. Click **Open Singapore CRM Report →** in the *Classify Leads* tab
2. Export as **CSV** or **Excel**
3. Required columns: `GRID`, `Account Name`, `Account Status`, `Restaurant PostalCode`, `Formatted Restaurant Address`

---

### Step 4 · Run Classification
1. Go to the **📊 Classify Leads** tab
2. Upload all three files:
   - **Leads** — from Step 1
   - **Apify Results with GRID** — from Step 2B
   - **CRM Export** — from Step 3
3. Click **▶ Run Classification**
4. Download the Excel report (filename: `ALG_DDMMYY`)

---

### Agent Columns (A–F in the Excel Report)

The Classified Leads sheet has 6 editable columns for your agents to fill in:

| Column | Purpose |
|---|---|
| **Agent** | Assign the lead to an agent |
| **Due Date** | Set a follow-up date |
| **Convert/Lost** | Dropdown: `Converted` or `Lost` |
| **Invalid Reason** | Dropdown: `Duplicate`, `Invalid Data`, `Closed Down`, `Wrong Target Group`, `Other` |
| **Comments/Duplicate GRID** | Free text — notes or the duplicate GRID reference |
| **Feedback** | Free text — any additional context |

---

### Classification Labels

| Label | What it means | What to do |
|---|---|---|
| **✅ P1 — New** | No match in CRM. Google confirms it's an open restaurant. | Pitch delivery to this lead |
| **🟡 P3 — Potential Match** | Name is 50–74% similar to a CRM account at the same postal code. | Verify manually — could be a duplicate or a different restaurant |
| **🔴 P4 — Duplicate** | Name is ≥ 75% similar to a CRM account at the same postal code. | Skip — already in the system |
| **🏢 Business Closed** | Google Maps confirms the restaurant is permanently or temporarily closed. | Skip |
| **❌ Wrong Target Group** | Google Maps category is not food delivery eligible. | Skip |
| **⚪ P2 — Please Check** | No Google Maps result found, or the result doesn't match the lead. | Verify manually before acting |

---

### Matching Logic

1. **Postal code** — only CRM accounts at the same 6-digit postal are compared
2. **Unit number** — if the lead has a unit (e.g. #01-23), all CRM accounts at that unit are scored
3. **Name similarity** — fuzzy matching + Hanyu Pinyin for Chinese names
   - Score ≥ 75% → **P4 Duplicate**
   - Score 50–74% → **P3 Potential Match**
   - Score < 50% → Apify check

---

### SF Account Audit (Tab 3)
Find suspected duplicate accounts within Salesforce itself. Run monthly for CRM hygiene.

---

### Tips
- Thresholds (P3 / P4) are adjustable in the sidebar
- Chinese names are auto-converted to Hanyu Pinyin
- Exclusion keywords control which Google Maps categories become Wrong Target Group
- `mix & match` and hawker stalls are handled correctly by default
        """)



if __name__ == "__main__":
    main()
